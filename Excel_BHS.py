from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from openpyxl.styles import Alignment


input_file = r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Data input\6.BHS.xlsx'
input_sheet = 'Sheet1'
output_path = r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Python output\2148BHS_24oct.xlsx'

df = pd.read_excel(input_file,input_sheet)
df=df[df['Q2'] == 1]








# Writing Error index in dataframe

# Error_n = ['Count1: Zone Quota ']
# df_name= pd.DataFrame(Error_n) 
# df_name.to_excel(output_path, index=False, header=False)


# Step 1: Create a list of topic names
topics = ['Count1: Zone Quota', 'Count2: SEC count', 'Count3 : Age count', 'Count4 : MOUB count', 'Count5:Age wise brands', 'Count6:SEC wise brands','Count7:SEC wise smoking category','Error1: Time outside of 8:00-20:00', 'Error2: Age vs CWE', ' Error3:No. of categories coded for Q11 ','Error4:No. of sticks', 'Error5: Started smoking before 15 or after 25 years age ', ' Error6:frequency of number of brands selected  by each interviewer in spontaneous awareness(Q14)','Error7:Regular brand not in TOM ','Error8:Regular brand not in TOM  or spont','Error9: Frequency of number of brands selected  by each interviewer in aided awareness(Q15)','Error10: Frequency of number of brands selected  by each interviewer in ever smoked(Q16)','Error11: Frequency of number of brands selected  by each interviewer in ever smoked in last 2 years (Q17)','Error12: Frequency of number of brands selected  by each interviewer in ever smoked in last 1 year (Q18)','Error13: Frequency of number of brands selected  by each interviewer in ever smoked in last 1 month (Q19)','Error14 Frequency of number of brands selected  by each interviewer in ever smoked in last 2 weeks(Q20.1)','Error15: Frequency of number of brands selected  by each interviewer in ever smoked in last 1 week(Q20)','Error16: No. of sticksof regular brand (Q23)','Error17: Difference between avg sticks smoked/day & regular brand stick smoked/day','Error18: Price fluctuation between Previous  & regular brand ','Error21: Percentage of loose vs pack','Error22: Sticks, a pack consist of?',' Error23: Mild or strong? (Q31)',' Error24: Regualr brand is Mild or strong? (Q32)',' Error26: Number of times Statements being coded (C1)','  Error27: Statement selection (C2)',' Error28: Just Right scale (C3)',' Error30: Reasons to smoke regular brand (D1)',' Error: Awareness of Cig with Capsule (H1)',' Error: Capsule bursting (H5)',' Error: Reasons not regularly smoke a Capsule cigarette (H7)',' Error: Seen this PACKET of cigarette brand Blue Charms? (S3)',' Error: How much do you like the PACKET of this cigarette brand? (S4)','  Error: how much you agree or disagree with each of the statement (S5)','  Error: GPS data absent','']


df_name = pd.DataFrame(topics, columns=['Logic list'])
df_name.to_excel(output_path, index=False)




def inside_append_dataframe_with_blank_rows(file_path, dataframe, blank_rows=2):
    """
    Appends a DataFrame to an existing Excel file with a specified number of blank rows in between.

    Parameters:
    - file_path: str, path to the Excel file
    - dataframe: pd.DataFrame, DataFrame to append
    - blank_rows: int, number of blank rows between appended DataFrames (default is 5)
    """
    
    with pd.ExcelWriter(
            file_path,
            engine='openpyxl',
            mode='a',
            if_sheet_exists='overlay') as writer:
        reader = pd.read_excel(file_path)
        dataframe.to_excel(
            writer,
            startrow=reader.shape[0] + blank_rows,
            index=True,
            header=True)

         
    # workbook = load_workbook(file_path)
    # worksheet = workbook.active

    # # Center-align the text in the appended DataFrame
    # for row in worksheet.iter_rows(min_row=startrow + 1, max_row=startrow + len(dataframe) + 3, min_col=1, max_col=len(dataframe.columns) + 1):
    #     for cell in row:
    #         cell.alignment = Alignment(horizontal='center', vertical='center')

    # # Save the workbook
    # workbook.save(file_path)     

 





def outside_append_dataframe_with_blank_rows(file_path, dataframe, blank_rows=5):
    """
    Appends a DataFrame to an existing Excel file with a specified number of blank rows in between.

    Parameters:
    - file_path: str, path to the Excel file
    - dataframe: pd.DataFrame, DataFrame to append
    - blank_rows: int, number of blank rows between appended DataFrames (default is 5)
    """
    
    with pd.ExcelWriter(
            file_path,
            engine='openpyxl',
            mode='a',
            if_sheet_exists='overlay') as writer:
        reader = pd.read_excel(file_path)
        dataframe.to_excel(
            writer,
            startrow=reader.shape[0] + blank_rows,
            index=True,
            header=True)











# Count1: Zone count  


Error_n = ['Count1 :Zone count']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()
input_string1=[1,2,3,4]
input_string2=['North','East','West','South']
dictionary = dict(zip(input_string1, input_string2))
df_global['Zone'] = df_global['Zone'].replace(dictionary)
df_s = df_global['Zone'].value_counts()
Valuecount=pd.DataFrame(df_s)
total_sum = Valuecount.sum()
total_sum.name = 'Total'
Valuecount=Valuecount._append(total_sum)
inside_append_dataframe_with_blank_rows(output_path, Valuecount)





# Count2: SEC count  


Error_n = ['Count2: SEC count ']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()
input_string1=[1,2,3,4]
input_string2=['SEC A','SEC B','SEC C','SEC D']
dictionary = dict(zip(input_string1, input_string2))
df_global['SEC_Q'] = df_global['SEC_Q'].replace(dictionary)
df_s = df_global['SEC_Q'].value_counts()
Valuecount=pd.DataFrame(df_s)
total_sum = Valuecount.sum()
total_sum.name = 'Total'
Valuecount=Valuecount._append(total_sum)
inside_append_dataframe_with_blank_rows(output_path, Valuecount)





# Count3 : Age count

Error_n = ['Count3 : Age count']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()
df_global.reset_index(drop=True, inplace=True)
df_global['Q6'] = pd.to_numeric(df_global['Q6'], errors='coerce')
def categorize_scores(score):
    
    if 21<=score <= 25:
        return '21-25 Yrs'
    elif 26 <= score <= 30:
        return '26-30 Yrs'
    elif 31<=score <= 35:
        return '31-35 Yrs'
    elif 36<=score<=40:
        return '36-40 Yrs'
    elif 41<=score<=45:
        return '41-45 Yrs'
    elif 46<=score<=50:
        return '46-50 Yrs'
    else:
        return 'Other'
    
    
df_global['Age'] = df_global['Q6'].apply(categorize_scores)       
counts_p = df_global['Age'].value_counts()
Valuecount=pd.DataFrame(counts_p)
total_sum = Valuecount.sum()
total_sum.name = 'Total'
Valuecount=Valuecount._append(total_sum)
inside_append_dataframe_with_blank_rows(output_path, Valuecount)







# Count4 : MOUB count

Error_n = ['Count4 : MOUB count ']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)


df_global=df.copy()
input_string1=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,63,64,65,66,67,68,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101]
string="American Club ,Cavanders Gold Rich Taste ,Chancellor ,Charminar Filter ,Charms Special Blue ,Classic ,Editions Trio ,Flake Special Filter ,Focus Mint  ,Gold Flake Kings ,Gold Flake Premium ,Gold Flake Premium Neo Smart ,Gold Flake Indie Mint ,Gold Flake Special ,Gold Flake Special Mint ,Gold Flake Super Star ,India King ,Marlboro KSFT ,Marlboro Advance Compact ,Marlboro Pocket Filter ,Navy Cut Fruit ,Navy Cut Virginia Filter ,Red & White Select ,Red & White Select,Regent ,Regent Black ,Regent Cool ,Silk Cut Blue ,Silk Cut Filter ,Silk Cut Virginia ,Special Blues ,Special Red Longs ,Special Red Signature ,Stellar Cool Blast  ,T3 White  ,Total Refresh ,Total Royal Twist ,Total Spearmint ,Wave Cool Mint ,Wills Navy Cut Filter ,Wills Navy Cut,American Club Mint ,American Club Smash ,American Fruit ,Benson & Hedges  ,Berkeley ,Berkely ,Blue Charms ,Blue Charms ,Bristol ,Cavander Gold  ,Chancellor ,Charminar Plains ,Charms Regular Filter ,Classic Connect ,Classic Ice Burst ,Club One ,Duke ,Editions Active Mint ,Editions Ice Fruit ,Editions Spark ,Flake Excel ,Flake Liberty ,Flake Mint Switch ,Flake Nova ,Flake White ,Gold Flake Century ,Gold Flake Filter ,Gold Flake Filter ,Gold Flake Kings ,Gold Flake Kings Lights ,Gold Flake Kings SLK ,Golden Gold Flake ,Kingston ,Marlboro Clove ,Marlboro Fuse Beyond ,NATIONAL GOLD FLAKE ,Navy Cut Deluxe Filter ,Originals ,Panama ,Panama Filter ,Panama Filter ,Player's Fruit ,Player's Mint ,Royal ,Scissors Menthol  ,Classic ,Stellar Slims Define ,Stellar Slims Shift ,Total Active Mint ,Wave Fruity ,Wave Mint ,Will Flake Premium Filter ,Zaffran ,American Club Clove Magik ,Classic AlphaTec ,Classic Double Burst ,Classic Verve ,Stellar Define Pan ,Stellar Shift Duos ,Wills Insignia "
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
dictionary = dict(zip(input_string1, input_string2))
df_global['Q21'] = df_global['Q21'].replace(dictionary)



df_modified = df_global.loc[:,['Q21','Interviewer']]
crosstab_result = pd.crosstab(df_modified['Q21'], df_modified['Interviewer'])
count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)

# total_number = 200


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total
# count2=(count1 / total_number) * 100



total_interviews = crosstab_result.values.sum()
crosstab_percentage = (crosstab_result / total_interviews) * 100
crosstab_percentage = crosstab_percentage.round(0)
crosstab_percentage['(%)'] = crosstab_percentage.sum(axis=1)
crosstab_percentage = crosstab_percentage[['(%)'] + [col for col in crosstab_percentage.columns if col != '(%)']]

count2=pd.DataFrame(crosstab_percentage )
# count2['Total']=count1.sum(min_count=1,axis=1)
# total=count2.sum()
# count2.loc['Total']=total




inside_append_dataframe_with_blank_rows(output_path, count1)
inside_append_dataframe_with_blank_rows(output_path, count2)

# counts_p = df['Q21'].value_counts()
# Valuecount=pd.DataFrame(counts_p)
# total_sum = Valuecount.sum()
# total_sum.name = 'Total'
# Valuecount=Valuecount._append(total_sum)
# inside_append_dataframe_with_blank_rows(output_path, Valuecount)









# Count5:Age wise brands 


Error_n = ['Count5:Age wise brands']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()



input_string1=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,63,64,65,66,67,68,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101]
string="American Club ,Cavanders Gold Rich Taste ,Chancellor ,Charminar Filter ,Charms Special Blue ,Classic ,Editions Trio ,Flake Special Filter ,Focus Mint  ,Gold Flake Kings ,Gold Flake Premium ,Gold Flake Premium Neo Smart ,Gold Flake Indie Mint ,Gold Flake Special ,Gold Flake Special Mint ,Gold Flake Super Star ,India King ,Marlboro KSFT ,Marlboro Advance Compact ,Marlboro Pocket Filter ,Navy Cut Fruit ,Navy Cut Virginia Filter ,Red & White Select ,Red & White Select,Regent ,Regent Black ,Regent Cool ,Silk Cut Blue ,Silk Cut Filter ,Silk Cut Virginia ,Special Blues ,Special Red Longs ,Special Red Signature ,Stellar Cool Blast  ,T3 White  ,Total Refresh ,Total Royal Twist ,Total Spearmint ,Wave Cool Mint ,Wills Navy Cut Filter ,Wills Navy Cut,American Club Mint ,American Club Smash ,American Fruit ,Benson & Hedges  ,Berkeley ,Berkely ,Blue Charms ,Blue Charms ,Bristol ,Cavander Gold  ,Chancellor ,Charminar Plains ,Charms Regular Filter ,Classic Connect ,Classic Ice Burst ,Club One ,Duke ,Editions Active Mint ,Editions Ice Fruit ,Editions Spark ,Flake Excel ,Flake Liberty ,Flake Mint Switch ,Flake Nova ,Flake White ,Gold Flake Century ,Gold Flake Filter ,Gold Flake Filter ,Gold Flake Kings ,Gold Flake Kings Lights ,Gold Flake Kings SLK ,Golden Gold Flake ,Kingston ,Marlboro Clove ,Marlboro Fuse Beyond ,NATIONAL GOLD FLAKE ,Navy Cut Deluxe Filter ,Originals ,Panama ,Panama Filter ,Panama Filter ,Player's Fruit ,Player's Mint ,Royal ,Scissors Menthol  ,Classic ,Stellar Slims Define ,Stellar Slims Shift ,Total Active Mint ,Wave Fruity ,Wave Mint ,Will Flake Premium Filter ,Zaffran ,American Club Clove Magik ,Classic AlphaTec ,Classic Double Burst ,Classic Verve ,Stellar Define Pan ,Stellar Shift Duos ,Wills Insignia "
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
dictionary = dict(zip(input_string1, input_string2))
df_global['Q21'] = df_global['Q21'].replace(dictionary)



df_global.reset_index(drop=True, inplace=True)
df_global['Q6'] = pd.to_numeric(df_global['Q6'], errors='coerce')
def categorize_scores(score):
    
    if 21<=score <= 25:
        return '21-25 Yrs'
    elif 26 <= score <= 30:
        return '26-30 Yrs'
    elif 31<=score <= 35:
        return '31-35 Yrs'
    elif 36<=score<=40:
        return '36-40 Yrs'
    elif 41<=score<=45:
        return '41-45 Yrs'
    elif 46<=score<=50:
        return '46-50 Yrs'
    else:
        return 'Other'
    
df_global['Age'] = df_global['Q6'].apply(categorize_scores)  


df_modified = df_global.loc[:,['Age','Q21']]
crosstab_result = pd.crosstab(df_modified['Age'], df_modified['Q21'])
count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


inside_append_dataframe_with_blank_rows(output_path, count1)






# Count6:SEC wise brands


Error_n = ['Count6:SEC wise brands']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()

input_string1=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,63,64,65,66,67,68,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101]
string="American Club ,Cavanders Gold Rich Taste ,Chancellor ,Charminar Filter ,Charms Special Blue ,Classic ,Editions Trio ,Flake Special Filter ,Focus Mint  ,Gold Flake Kings ,Gold Flake Premium ,Gold Flake Premium Neo Smart ,Gold Flake Indie Mint ,Gold Flake Special ,Gold Flake Special Mint ,Gold Flake Super Star ,India King ,Marlboro KSFT ,Marlboro Advance Compact ,Marlboro Pocket Filter ,Navy Cut Fruit ,Navy Cut Virginia Filter ,Red & White Select ,Red & White Select,Regent ,Regent Black ,Regent Cool ,Silk Cut Blue ,Silk Cut Filter ,Silk Cut Virginia ,Special Blues ,Special Red Longs ,Special Red Signature ,Stellar Cool Blast  ,T3 White  ,Total Refresh ,Total Royal Twist ,Total Spearmint ,Wave Cool Mint ,Wills Navy Cut Filter ,Wills Navy Cut,American Club Mint ,American Club Smash ,American Fruit ,Benson & Hedges  ,Berkeley ,Berkely ,Blue Charms ,Blue Charms ,Bristol ,Cavander Gold  ,Chancellor ,Charminar Plains ,Charms Regular Filter ,Classic Connect ,Classic Ice Burst ,Club One ,Duke ,Editions Active Mint ,Editions Ice Fruit ,Editions Spark ,Flake Excel ,Flake Liberty ,Flake Mint Switch ,Flake Nova ,Flake White ,Gold Flake Century ,Gold Flake Filter ,Gold Flake Filter ,Gold Flake Kings ,Gold Flake Kings Lights ,Gold Flake Kings SLK ,Golden Gold Flake ,Kingston ,Marlboro Clove ,Marlboro Fuse Beyond ,NATIONAL GOLD FLAKE ,Navy Cut Deluxe Filter ,Originals ,Panama ,Panama Filter ,Panama Filter ,Player's Fruit ,Player's Mint ,Royal ,Scissors Menthol  ,Classic ,Stellar Slims Define ,Stellar Slims Shift ,Total Active Mint ,Wave Fruity ,Wave Mint ,Will Flake Premium Filter ,Zaffran ,American Club Clove Magik ,Classic AlphaTec ,Classic Double Burst ,Classic Verve ,Stellar Define Pan ,Stellar Shift Duos ,Wills Insignia "
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
dictionary = dict(zip(input_string1, input_string2))
df_global['Q21'] = df_global['Q21'].replace(dictionary)

input_string1=[1,2,3,4]
input_string2=['SEC A','SEC B','SEC C','SEC D']
dictionary = dict(zip(input_string1, input_string2))
df_global['SEC_Q'] = df_global['SEC_Q'].replace(dictionary)


df_modified = df_global.loc[:,['SEC_Q','Q21']]
crosstab_result = pd.crosstab(df_modified['SEC_Q'], df_modified['Q21'])
count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


inside_append_dataframe_with_blank_rows(output_path, count1)





# Count7:SEC wise smoking category  


Error_n = ['Count7:SEC wise smoking category ']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()
input_string1=[1,2,3,4]
input_string2=['SEC A','SEC B','SEC C','SEC D']
dictionary = dict(zip(input_string1, input_string2))
df_global['SEC_Q'] = df_global['SEC_Q'].replace(dictionary)
# df_s = df_global['SEC_Q'].value_counts()
# Valuecount=pd.DataFrame(df_s)
# total_sum = Valuecount.sum()
# total_sum.name = 'Total'
# Valuecount=Valuecount._append(total_sum)
# inside_append_dataframe_with_blank_rows(output_path, Valuecount)




result = df_global.groupby('SEC_Q')[['A_Q11_1', 'A_Q11_2', 'A_Q11_3','A_Q11_4','A_Q11_5','A_Q11_6','A_Q11_7','A_Q11_8','A_Q11_9']].sum()
result = result.rename(columns={
    'A_Q11_1': 'Beedi Smokers',
    'A_Q11_2': 'Cigar/Cheroot',
    'A_Q11_3': 'Cigarette',
    'A_Q11_4': 'E-cigarette',
    'A_Q11_5': 'Gutkha (Packaged or Loose) / Mawa',
    'A_Q11_6': 'Khaini (Packaged or Loose)',
    'A_Q11_7': 'Pan Masala (Sada)',
    'A_Q11_8': 'Tobacco chewing gum',
    'A_Q11_9': 'Zarda with pan'
    
})

# total=result.sum()
# total.name='Total'
# result=result._append(total)
# inside_append_dataframe_with_blank_rows(output_path, result)



# crosstab_result = pd.crosstab(df_global['SEC_Q'], result)



# dynamic_headers = [f'brands{i}' for i in range(1, len(crosstab_result.columns) + 1)]

# count1=pd.DataFrame(crosstab_result )
# count1['Total']=count1.sum(min_count=1,axis=1)


# total=count1.sum()
# # total.name='Total'
# # count1=count1._append(total)
# count1.loc['Total']=total


# inside_append_dataframe_with_blank_rows(output_path, df_modified)
inside_append_dataframe_with_blank_rows(output_path, result)


















# Error1: Time outside of 8:00-20:00
Error_n = ['Error1: Time outside of 8:00-20:00']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)






df_global=df.copy()
df_global['Date'] = pd.to_datetime(df_global['Date'], format='%d-%m-%Y %H:%M:%S')

# Filter rows where time is less than 08:00:00 or greater than 22:00:00
filtered_df = df_global[(df_global['Date'].dt.time < pd.to_datetime('08:00:00').time()) | (df_global['Date'].dt.time > pd.to_datetime('22:00:00').time())]
filtered_df=filtered_df[['Interviewer', 'Date']]



counts_p = filtered_df['Interviewer'].value_counts()
Valuecount=pd.DataFrame(counts_p)
total_sum = Valuecount.sum()
total_sum.name = 'Total'
Valuecount=Valuecount._append(total_sum)
inside_append_dataframe_with_blank_rows(output_path, Valuecount)












# Error2: Age vs CWE

Error_n = ['Error2: Age vs CWE']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()
df_global['Q6'] = pd.to_numeric(df_global['Q6'], errors='coerce')
def categorize_scores(score):
    
    if 21<=score <= 25:
        return '21-25 Yrs'
    elif 26 <= score <= 30:
        return '26-30 Yrs'
    elif 31<=score <= 35:
        return '31-35 Yrs'
    elif 36<=score<=40:
        return '36-40 Yrs'
    elif 41<=score<=45:
        return '41-45 Yrs'
    elif 46<=score<=50:
        return '46-50 Yrs'
    else:
        return 'Other'
    
df_global['Age'] = df_global['Q6'].apply(categorize_scores)  

input_string1=[1,2]
input_string2=['Yes','No']
dictionary = dict(zip(input_string1, input_string2))
df_global['CWE'] = df_global['Q6_1'].replace(dictionary)


df_modified = df_global.loc[:,['Interviewer','Age','CWE']]


crosstab = pd.crosstab(df_modified['Interviewer'], df_modified['Age'], values=df_modified['CWE'],aggfunc=lambda x: (x == 'No').sum(), 
    margins=True, 
    margins_name='Total')


inside_append_dataframe_with_blank_rows(output_path, crosstab)










# Error3:No. of categories coded for Q11

Error_n = [' Error3:No. of categories coded for Q11']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()

result = df_global.groupby('Interviewer')[['A_Q11_1', 'A_Q11_2', 'A_Q11_3','A_Q11_4','A_Q11_5','A_Q11_6','A_Q11_7','A_Q11_8','A_Q11_9']].sum()
result = result.rename(columns={
    'A_Q11_1': 'Beedi Smokers',
    'A_Q11_2': 'Cigar/Cheroot',
    'A_Q11_3': 'Cigarette',
    'A_Q11_4': 'E-cigarette',
    'A_Q11_5': 'Gutkha (Packaged or Loose) / Mawa',
    'A_Q11_6': 'Khaini (Packaged or Loose)',
    'A_Q11_7': 'Pan Masala (Sada)',
    'A_Q11_8': 'Tobacco chewing gum',
    'A_Q11_9': 'Zarda with pan'
    
})

total=result.sum()
total.name='Total'
result=result._append(total)
inside_append_dataframe_with_blank_rows(output_path, result)












# Error4:No. of sticks

Error_n = [' Error4:No. of sticks']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()



df_modified = df_global.loc[:,['Q11_2','Interviewer']]
crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['Q11_2'])
count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


inside_append_dataframe_with_blank_rows(output_path, count1)








# Error5: Started smoking before 15 or after 25 years age

Error_n = [' Error5: Started smoking before 15 or after 25 years age']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()
df_global['Q12'] = pd.to_numeric(df_global['Q12'], errors='coerce')


# Filter rows where column 'A' is less than 5 or greater than 20
filtered_df = df_global[(df_global['Q12'] < 15) | (df_global['Q12'] > 25)]

# Select columns 'A', 'B', and 'C'
result_df = filtered_df[['Interviewer', 'Respondent_id', 'Q12']]


result_df.rename(columns={
    'Q12': 'Smoking age',
    'Respondent_id': 'Respondent',
    'Interviewer': 'Interviewer'
    }, inplace=True)


# filt = ((df_global['Q12']>25) & (df_global['Q12']<15))

                                              
df_s = result_df['Interviewer'].value_counts()

Valuecount=pd.DataFrame(df_s)
total_sum = Valuecount.sum()
total_sum.name = 'Total'
Valuecount=Valuecount._append(total_sum)
# inside_append_dataframe_with_blank_rows(output_path, df_modified)
inside_append_dataframe_with_blank_rows(output_path, Valuecount)
inside_append_dataframe_with_blank_rows(output_path, result_df)











# Error6:frequency of number of brands selected  by each interviewer in spontaneous awareness(Q14)

Error_n = [' Error6:frequency of number of brands selected  by each interviewer in spontaneous awareness(Q14)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()


string="A_Q14_1_1,A_Q14_1_2,A_Q14_1_3,A_Q14_1_4,A_Q14_1_5,A_Q14_1_6,A_Q14_1_7,A_Q14_1_8,A_Q14_1_9,A_Q14_1_10,A_Q14_1_11,A_Q14_1_12,A_Q14_1_13,A_Q14_1_14,A_Q14_1_15,A_Q14_1_16,A_Q14_1_17,A_Q14_1_18,A_Q14_1_19,A_Q14_1_20,A_Q14_1_21,A_Q14_1_22,A_Q14_1_23,A_Q14_1_24,A_Q14_1_25,A_Q14_1_26,A_Q14_1_27,A_Q14_1_28,A_Q14_1_29,A_Q14_1_30,A_Q14_1_31,A_Q14_1_32,A_Q14_1_33,A_Q14_1_34,A_Q14_1_35,A_Q14_1_36,A_Q14_1_37,A_Q14_1_38,A_Q14_1_39,A_Q14_1_40,A_Q14_1_41,A_Q14_1_42,A_Q14_1_43,A_Q14_1_44,A_Q14_1_45,A_Q14_1_46,A_Q14_1_47,A_Q14_1_48,A_Q14_1_49,A_Q14_1_50,A_Q14_1_51,A_Q14_1_52,A_Q14_1_53,A_Q14_1_54,A_Q14_1_55,A_Q14_1_56,A_Q14_1_57,A_Q14_1_58,A_Q14_1_59,A_Q14_1_60,A_Q14_1_61,A_Q14_1_62,A_Q14_1_63,A_Q14_1_64,A_Q14_1_65,A_Q14_1_66,A_Q14_1_67,A_Q14_1_68,A_Q14_1_69,A_Q14_1_70,A_Q14_1_71,A_Q14_1_72,A_Q14_1_73,A_Q14_1_74,A_Q14_1_75,A_Q14_1_76,A_Q14_1_77,A_Q14_1_78,A_Q14_1_79,A_Q14_1_80,A_Q14_1_81,A_Q14_1_82,A_Q14_1_83,A_Q14_1_84,A_Q14_1_85,A_Q14_1_86,A_Q14_1_87,A_Q14_1_88,A_Q14_1_89,A_Q14_1_90,A_Q14_1_91,A_Q14_1_92,A_Q14_1_93,A_Q14_1_94,A_Q14_1_95,A_Q14_1_96,A_Q14_1_97,A_Q14_1_98,A_Q14_1_99,A_Q14_1_100,A_Q14_1_101"
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
df_global['sum_columns'] = df_global[input_string2].sum(axis=1)
df_modified = df_global.loc[:,['Respondent_id','Interviewer','sum_columns']]

crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['sum_columns'])



# dynamic_headers = [f'brands{i}' for i in range(1, len(crosstab_result.columns) + 1)]

count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


inside_append_dataframe_with_blank_rows(output_path, count1)














# Error7:Regular brand not in TOM

Error_n = [' Error7:Regular brand not in TOM']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()

filt = (df_global['Q14'] != df_global['Q21']) 

nom = df_global.loc[filt,['Respondent_id','Interviewer','Q14','Q21']]
df_modified = nom                                         
df_s = df_modified['Interviewer'].value_counts()

Valuecount=pd.DataFrame(df_s)
total_sum = Valuecount.sum()
total_sum.name = 'Total'
Valuecount=Valuecount._append(total_sum)
inside_append_dataframe_with_blank_rows(output_path, df_modified)
inside_append_dataframe_with_blank_rows(output_path, Valuecount)











# Error8:Regular brand not in TOM  or spont

Error_n = [' Error8:Regular brand not in TOM  or spont']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()
input_string1=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,63,64,65,66,67,68,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101]
string="A_Q14_1_1,A_Q14_1_2,A_Q14_1_3,A_Q14_1_4,A_Q14_1_5,A_Q14_1_6,A_Q14_1_7,A_Q14_1_8,A_Q14_1_9,A_Q14_1_10,A_Q14_1_11,A_Q14_1_12,A_Q14_1_13,A_Q14_1_14,A_Q14_1_15,A_Q14_1_16,A_Q14_1_17,A_Q14_1_18,A_Q14_1_19,A_Q14_1_20,A_Q14_1_21,A_Q14_1_22,A_Q14_1_23,A_Q14_1_24,A_Q14_1_25,A_Q14_1_26,A_Q14_1_27,A_Q14_1_28,A_Q14_1_29,A_Q14_1_30,A_Q14_1_31,A_Q14_1_32,A_Q14_1_33,A_Q14_1_34,A_Q14_1_35,A_Q14_1_36,A_Q14_1_37,A_Q14_1_38,A_Q14_1_39,A_Q14_1_40,A_Q14_1_41,A_Q14_1_42,A_Q14_1_43,A_Q14_1_44,A_Q14_1_45,A_Q14_1_46,A_Q14_1_47,A_Q14_1_48,A_Q14_1_49,A_Q14_1_50,A_Q14_1_51,A_Q14_1_52,A_Q14_1_53,A_Q14_1_54,A_Q14_1_55,A_Q14_1_56,A_Q14_1_57,A_Q14_1_58,A_Q14_1_59,A_Q14_1_60,A_Q14_1_61,A_Q14_1_62,A_Q14_1_63,A_Q14_1_64,A_Q14_1_65,A_Q14_1_66,A_Q14_1_67,A_Q14_1_68,A_Q14_1_69,A_Q14_1_70,A_Q14_1_71,A_Q14_1_72,A_Q14_1_73,A_Q14_1_74,A_Q14_1_75,A_Q14_1_76,A_Q14_1_77,A_Q14_1_78,A_Q14_1_79,A_Q14_1_80,A_Q14_1_81,A_Q14_1_82,A_Q14_1_83,A_Q14_1_84,A_Q14_1_85,A_Q14_1_86,A_Q14_1_87,A_Q14_1_88,A_Q14_1_89,A_Q14_1_90,A_Q14_1_91,A_Q14_1_92,A_Q14_1_93,A_Q14_1_94,A_Q14_1_95,A_Q14_1_96,A_Q14_1_97,A_Q14_1_98,A_Q14_1_99,A_Q14_1_100,A_Q14_1_101"
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
dictionary = dict(zip(input_string1, input_string2))


for df_row in range(len(df_global)):
        if (df_global.loc[df_row,dictionary[df_global.loc[df_row,'Q21']]]== 1  or  df_global.loc[df_row,'Q14']==df_global.loc[df_row,'Q21']):
            df_global.drop(df_row,inplace=True)
        else:
             pass
         

additional_columns = ['Respondent_id', 'Name','Interviewer','Q21']

# Combine the lists
all_columns = additional_columns+input_string2


df_modified = df_global.loc[:,all_columns]




                                                  
df_s = df_modified['Interviewer'].value_counts()

Valuecount=pd.DataFrame(df_s)
total_sum = Valuecount.sum()
total_sum.name = 'Total'
Valuecount=Valuecount._append(total_sum)
inside_append_dataframe_with_blank_rows(output_path, df_modified)
inside_append_dataframe_with_blank_rows(output_path, Valuecount)










# Error9: Frequency of number of brands selected  by each interviewer in aided awareness(Q15)

Error_n = ['Error9: Frequency of number of brands selected  by each interviewer in aided awareness(Q15)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()
df_global['sum_columns'] = df_global[['A_Q15_1','A_Q15_2','A_Q15_3','A_Q15_4','A_Q15_5','A_Q15_6','A_Q15_7','A_Q15_8','A_Q15_9','A_Q15_10','A_Q15_11','A_Q15_12','A_Q15_13','A_Q15_14','A_Q15_15','A_Q15_16','A_Q15_17','A_Q15_18','A_Q15_19','A_Q15_20','A_Q15_21','A_Q15_22','A_Q15_23','A_Q15_24','A_Q15_25','A_Q15_26','A_Q15_27','A_Q15_28','A_Q15_29','A_Q15_30','A_Q15_31','A_Q15_32','A_Q15_33','A_Q15_34','A_Q15_35','A_Q15_36','A_Q15_37','A_Q15_38','A_Q15_39','A_Q15_40','A_Q15_41','A_Q15_42','A_Q15_43','A_Q15_44','A_Q15_45','A_Q15_46','A_Q15_47','A_Q15_48','A_Q15_49','A_Q15_50','A_Q15_51','A_Q15_52','A_Q15_53','A_Q15_54','A_Q15_55','A_Q15_56','A_Q15_57','A_Q15_58','A_Q15_59','A_Q15_60','A_Q15_61','A_Q15_62','A_Q15_63','A_Q15_64','A_Q15_65','A_Q15_66','A_Q15_67','A_Q15_68','A_Q15_69','A_Q15_70','A_Q15_71','A_Q15_72','A_Q15_73','A_Q15_74','A_Q15_75','A_Q15_76','A_Q15_77','A_Q15_78','A_Q15_79','A_Q15_80','A_Q15_81','A_Q15_82','A_Q15_83','A_Q15_84','A_Q15_85','A_Q15_86','A_Q15_87','A_Q15_88','A_Q15_89','A_Q15_90','A_Q15_91','A_Q15_92','A_Q15_93','A_Q15_94','A_Q15_95','A_Q15_96','A_Q15_97','A_Q15_98','A_Q15_99','A_Q15_100','A_Q15_101']].sum(axis=1)
df_modified = df_global.loc[:,['Respondent_id','Interviewer','sum_columns']]

crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['sum_columns'])
crosstab_result['Total']=crosstab_result.sum(min_count=1,axis=1)
total=crosstab_result.sum()
total.name='Total'
crosstab_result=crosstab_result._append(total)
count1=pd.DataFrame(crosstab_result )
inside_append_dataframe_with_blank_rows(output_path, df_modified)
inside_append_dataframe_with_blank_rows(output_path, count1)













# Error10: Frequency of number of brands selected  by each interviewer in ever smoked(Q16)

Error_n = ['Error10: Frequency of number of brands selected  by each interviewer in ever smoked(Q16)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()
df_global['sum_columns'] = df_global[['A_Q16_1','A_Q16_2','A_Q16_3','A_Q16_4','A_Q16_5','A_Q16_6','A_Q16_7','A_Q16_8','A_Q16_9','A_Q16_10','A_Q16_11','A_Q16_12','A_Q16_13','A_Q16_14','A_Q16_15','A_Q16_16','A_Q16_17','A_Q16_18','A_Q16_19','A_Q16_20','A_Q16_21','A_Q16_22','A_Q16_23','A_Q16_24','A_Q16_25','A_Q16_26','A_Q16_27','A_Q16_28','A_Q16_29','A_Q16_30','A_Q16_31','A_Q16_32','A_Q16_33','A_Q16_34','A_Q16_35','A_Q16_36','A_Q16_37','A_Q16_38','A_Q16_39','A_Q16_40','A_Q16_41','A_Q16_42','A_Q16_43','A_Q16_44','A_Q16_45','A_Q16_46','A_Q16_47','A_Q16_48','A_Q16_49','A_Q16_50','A_Q16_51','A_Q16_52','A_Q16_53','A_Q16_54','A_Q16_55','A_Q16_56','A_Q16_57','A_Q16_58','A_Q16_59','A_Q16_60','A_Q16_61','A_Q16_62','A_Q16_63','A_Q16_64','A_Q16_65','A_Q16_66','A_Q16_67','A_Q16_68','A_Q16_69','A_Q16_70','A_Q16_71','A_Q16_72','A_Q16_73','A_Q16_74','A_Q16_75','A_Q16_76','A_Q16_77','A_Q16_78','A_Q16_79','A_Q16_80','A_Q16_81','A_Q16_82','A_Q16_83','A_Q16_84','A_Q16_85','A_Q16_86','A_Q16_87','A_Q16_88','A_Q16_89','A_Q16_90','A_Q16_91','A_Q16_92','A_Q16_93','A_Q16_94','A_Q16_95','A_Q16_96','A_Q16_97','A_Q16_98','A_Q16_99','A_Q16_100','A_Q16_101']].sum(axis=1)


df_modified = df_global.loc[:,['Respondent_id','Interviewer','sum_columns']]

crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['sum_columns'])
crosstab_result['Total']=crosstab_result.sum(min_count=1,axis=1)
total=crosstab_result.sum()
total.name='Total'
crosstab_result=crosstab_result._append(total)
count1=pd.DataFrame(crosstab_result )
inside_append_dataframe_with_blank_rows(output_path, df_modified)
inside_append_dataframe_with_blank_rows(output_path, count1)







# Error11: Frequency of number of brands selected  by each interviewer in ever smoked in last 2 years (Q17)

Error_n = ['Error11: Frequency of number of brands selected  by each interviewer in ever smoked in last 2 years (Q17)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()
string="A_Q17_1,A_Q17_2,A_Q17_3,A_Q17_4,A_Q17_5,A_Q17_6,A_Q17_7,A_Q17_8,A_Q17_9,A_Q17_10,A_Q17_11,A_Q17_12,A_Q17_13,A_Q17_14,A_Q17_15,A_Q17_16,A_Q17_17,A_Q17_18,A_Q17_19,A_Q17_20,A_Q17_21,A_Q17_22,A_Q17_23,A_Q17_24,A_Q17_25,A_Q17_26,A_Q17_27,A_Q17_28,A_Q17_29,A_Q17_30,A_Q17_31,A_Q17_32,A_Q17_33,A_Q17_34,A_Q17_35,A_Q17_36,A_Q17_37,A_Q17_38,A_Q17_39,A_Q17_40,A_Q17_41,A_Q17_42,A_Q17_43,A_Q17_44,A_Q17_45,A_Q17_46,A_Q17_47,A_Q17_48,A_Q17_49,A_Q17_50,A_Q17_51,A_Q17_52,A_Q17_53,A_Q17_54,A_Q17_55,A_Q17_56,A_Q17_57,A_Q17_58,A_Q17_59,A_Q17_60,A_Q17_61,A_Q17_62,A_Q17_63,A_Q17_64,A_Q17_65,A_Q17_66,A_Q17_67,A_Q17_68,A_Q17_69,A_Q17_70,A_Q17_71,A_Q17_72,A_Q17_73,A_Q17_74,A_Q17_75,A_Q17_76,A_Q17_77,A_Q17_78,A_Q17_79,A_Q17_80,A_Q17_81,A_Q17_82,A_Q17_83,A_Q17_84,A_Q17_85,A_Q17_86,A_Q17_87,A_Q17_88,A_Q17_89,A_Q17_90,A_Q17_91,A_Q17_92,A_Q17_93,A_Q17_94,A_Q17_95,A_Q17_96,A_Q17_97,A_Q17_98,A_Q17_99,A_Q17_100,A_Q17_101"
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
df_global['sum_columns'] = df_global[input_string2].sum(axis=1)
df_modified = df_global.loc[:,['Respondent_id','Interviewer','sum_columns']]

crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['sum_columns'])



# dynamic_headers = [f'brands{i}' for i in range(1, len(crosstab_result.columns) + 1)]

count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


# inside_append_dataframe_with_blank_rows(output_path, df_modified)
inside_append_dataframe_with_blank_rows(output_path, count1)













# Error12: Frequency of number of brands selected  by each interviewer in ever smoked in last 1 year (Q18)

Error_n = ['Error12: Frequency of number of brands selected  by each interviewer in ever smoked in last 1 year (Q18)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()
string="A_Q18_1,A_Q18_2,A_Q18_3,A_Q18_4,A_Q18_5,A_Q18_6,A_Q18_7,A_Q18_8,A_Q18_9,A_Q18_10,A_Q18_11,A_Q18_12,A_Q18_13,A_Q18_14,A_Q18_15,A_Q18_16,A_Q18_17,A_Q18_18,A_Q18_19,A_Q18_20,A_Q18_21,A_Q18_22,A_Q18_23,A_Q18_24,A_Q18_25,A_Q18_26,A_Q18_27,A_Q18_28,A_Q18_29,A_Q18_30,A_Q18_31,A_Q18_32,A_Q18_33,A_Q18_34,A_Q18_35,A_Q18_36,A_Q18_37,A_Q18_38,A_Q18_39,A_Q18_40,A_Q18_41,A_Q18_42,A_Q18_43,A_Q18_44,A_Q18_45,A_Q18_46,A_Q18_47,A_Q18_48,A_Q18_49,A_Q18_50,A_Q18_51,A_Q18_52,A_Q18_53,A_Q18_54,A_Q18_55,A_Q18_56,A_Q18_57,A_Q18_58,A_Q18_59,A_Q18_60,A_Q18_61,A_Q18_62,A_Q18_63,A_Q18_64,A_Q18_65,A_Q18_66,A_Q18_67,A_Q18_68,A_Q18_69,A_Q18_70,A_Q18_71,A_Q18_72,A_Q18_73,A_Q18_74,A_Q18_75,A_Q18_76,A_Q18_77,A_Q18_78,A_Q18_79,A_Q18_80,A_Q18_81,A_Q18_82,A_Q18_83,A_Q18_84,A_Q18_85,A_Q18_86,A_Q18_87,A_Q18_88,A_Q18_89,A_Q18_90,A_Q18_91,A_Q18_92,A_Q18_93,A_Q18_94,A_Q18_95,A_Q18_96,A_Q18_97,A_Q18_98,A_Q18_99,A_Q18_100,A_Q18_101"
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
df_global['sum_columns'] = df_global[input_string2].sum(axis=1)
df_modified = df_global.loc[:,['Respondent_id','Interviewer','sum_columns']]

crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['sum_columns'])



# dynamic_headers = [f'brands{i}' for i in range(1, len(crosstab_result.columns) + 1)]

count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


# inside_append_dataframe_with_blank_rows(output_path, df_modified)
inside_append_dataframe_with_blank_rows(output_path, count1)











# Error13: Frequency of number of brands selected  by each interviewer in ever smoked in last 1 month (Q19)

Error_n = ['Error13: Frequency of number of brands selected  by each interviewer in ever smoked in last 1 month (Q19)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()
string="A_Q19_1,A_Q19_2,A_Q19_3,A_Q19_4,A_Q19_5,A_Q19_6,A_Q19_7,A_Q19_8,A_Q19_9,A_Q19_10,A_Q19_11,A_Q19_12,A_Q19_13,A_Q19_14,A_Q19_15,A_Q19_16,A_Q19_17,A_Q19_18,A_Q19_19,A_Q19_20,A_Q19_21,A_Q19_22,A_Q19_23,A_Q19_24,A_Q19_25,A_Q19_26,A_Q19_27,A_Q19_28,A_Q19_29,A_Q19_30,A_Q19_31,A_Q19_32,A_Q19_33,A_Q19_34,A_Q19_35,A_Q19_36,A_Q19_37,A_Q19_38,A_Q19_39,A_Q19_40,A_Q19_41,A_Q19_42,A_Q19_43,A_Q19_44,A_Q19_45,A_Q19_46,A_Q19_47,A_Q19_48,A_Q19_49,A_Q19_50,A_Q19_51,A_Q19_52,A_Q19_53,A_Q19_54,A_Q19_55,A_Q19_56,A_Q19_57,A_Q19_58,A_Q19_59,A_Q19_60,A_Q19_61,A_Q19_62,A_Q19_63,A_Q19_64,A_Q19_65,A_Q19_66,A_Q19_67,A_Q19_68,A_Q19_69,A_Q19_70,A_Q19_71,A_Q19_72,A_Q19_73,A_Q19_74,A_Q19_75,A_Q19_76,A_Q19_77,A_Q19_78,A_Q19_79,A_Q19_80,A_Q19_81,A_Q19_82,A_Q19_83,A_Q19_84,A_Q19_85,A_Q19_86,A_Q19_87,A_Q19_88,A_Q19_89,A_Q19_90,A_Q19_91,A_Q19_92,A_Q19_93,A_Q19_94,A_Q19_95,A_Q19_96,A_Q19_97,A_Q19_98,A_Q19_99,A_Q19_100,A_Q19_101"
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
df_global['sum_columns'] = df_global[input_string2].sum(axis=1)
df_modified = df_global.loc[:,['Respondent_id','Interviewer','sum_columns']]

crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['sum_columns'])



# dynamic_headers = [f'brands{i}' for i in range(1, len(crosstab_result.columns) + 1)]

count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


# inside_append_dataframe_with_blank_rows(output_path, df_modified)
inside_append_dataframe_with_blank_rows(output_path, count1)











# Error13: Frequency of number of brands selected  by each interviewer in ever smoked in last 1 month (Q19)

Error_n = ['Error13: Frequency of number of brands selected  by each interviewer in ever smoked in last 1 month (Q19)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()
string="A_Q19_1,A_Q19_2,A_Q19_3,A_Q19_4,A_Q19_5,A_Q19_6,A_Q19_7,A_Q19_8,A_Q19_9,A_Q19_10,A_Q19_11,A_Q19_12,A_Q19_13,A_Q19_14,A_Q19_15,A_Q19_16,A_Q19_17,A_Q19_18,A_Q19_19,A_Q19_20,A_Q19_21,A_Q19_22,A_Q19_23,A_Q19_24,A_Q19_25,A_Q19_26,A_Q19_27,A_Q19_28,A_Q19_29,A_Q19_30,A_Q19_31,A_Q19_32,A_Q19_33,A_Q19_34,A_Q19_35,A_Q19_36,A_Q19_37,A_Q19_38,A_Q19_39,A_Q19_40,A_Q19_41,A_Q19_42,A_Q19_43,A_Q19_44,A_Q19_45,A_Q19_46,A_Q19_47,A_Q19_48,A_Q19_49,A_Q19_50,A_Q19_51,A_Q19_52,A_Q19_53,A_Q19_54,A_Q19_55,A_Q19_56,A_Q19_57,A_Q19_58,A_Q19_59,A_Q19_60,A_Q19_61,A_Q19_62,A_Q19_63,A_Q19_64,A_Q19_65,A_Q19_66,A_Q19_67,A_Q19_68,A_Q19_69,A_Q19_70,A_Q19_71,A_Q19_72,A_Q19_73,A_Q19_74,A_Q19_75,A_Q19_76,A_Q19_77,A_Q19_78,A_Q19_79,A_Q19_80,A_Q19_81,A_Q19_82,A_Q19_83,A_Q19_84,A_Q19_85,A_Q19_86,A_Q19_87,A_Q19_88,A_Q19_89,A_Q19_90,A_Q19_91,A_Q19_92,A_Q19_93,A_Q19_94,A_Q19_95,A_Q19_96,A_Q19_97,A_Q19_98,A_Q19_99,A_Q19_100,A_Q19_101"
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
df_global['sum_columns'] = df_global[input_string2].sum(axis=1)
df_modified = df_global.loc[:,['Respondent_id','Interviewer','sum_columns']]

crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['sum_columns'])



# dynamic_headers = [f'brands{i}' for i in range(1, len(crosstab_result.columns) + 1)]

count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


# inside_append_dataframe_with_blank_rows(output_path, df_modified)
inside_append_dataframe_with_blank_rows(output_path, count1)

















# Error14 Frequency of number of brands selected  by each interviewer in ever smoked in last 2 weeks(Q20.1)

Error_n = ['Error14 Frequency of number of brands selected  by each interviewer in ever smoked in last 2 weeks(Q20.1)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()
string="A_Q20_1_1,A_Q20_1_2,A_Q20_1_3,A_Q20_1_4,A_Q20_1_5,A_Q20_1_6,A_Q20_1_7,A_Q20_1_8,A_Q20_1_9,A_Q20_1_10,A_Q20_1_11,A_Q20_1_12,A_Q20_1_13,A_Q20_1_14,A_Q20_1_15,A_Q20_1_16,A_Q20_1_17,A_Q20_1_18,A_Q20_1_19,A_Q20_1_20,A_Q20_1_21,A_Q20_1_22,A_Q20_1_23,A_Q20_1_24,A_Q20_1_25,A_Q20_1_26,A_Q20_1_27,A_Q20_1_28,A_Q20_1_29,A_Q20_1_30,A_Q20_1_31,A_Q20_1_32,A_Q20_1_33,A_Q20_1_34,A_Q20_1_35,A_Q20_1_36,A_Q20_1_37,A_Q20_1_38,A_Q20_1_39,A_Q20_1_40,A_Q20_1_41,A_Q20_1_42,A_Q20_1_43,A_Q20_1_44,A_Q20_1_45,A_Q20_1_46,A_Q20_1_47,A_Q20_1_48,A_Q20_1_49,A_Q20_1_50,A_Q20_1_51,A_Q20_1_52,A_Q20_1_53,A_Q20_1_54,A_Q20_1_55,A_Q20_1_56,A_Q20_1_57,A_Q20_1_58,A_Q20_1_59,A_Q20_1_60,A_Q20_1_61,A_Q20_1_62,A_Q20_1_63,A_Q20_1_64,A_Q20_1_65,A_Q20_1_66,A_Q20_1_67,A_Q20_1_68,A_Q20_1_69,A_Q20_1_70,A_Q20_1_71,A_Q20_1_72,A_Q20_1_73,A_Q20_1_74,A_Q20_1_75,A_Q20_1_76,A_Q20_1_77,A_Q20_1_78,A_Q20_1_79,A_Q20_1_80,A_Q20_1_81,A_Q20_1_82,A_Q20_1_83,A_Q20_1_84,A_Q20_1_85,A_Q20_1_86,A_Q20_1_87,A_Q20_1_88,A_Q20_1_89,A_Q20_1_90,A_Q20_1_91,A_Q20_1_92,A_Q20_1_93,A_Q20_1_94,A_Q20_1_95,A_Q20_1_96,A_Q20_1_97,A_Q20_1_98,A_Q20_1_99,A_Q20_1_100,A_Q20_1_101"
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
df_global['sum_columns'] = df_global[input_string2].sum(axis=1)
df_modified = df_global.loc[:,['Respondent_id','Interviewer','sum_columns']]

crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['sum_columns'])



# dynamic_headers = [f'brands{i}' for i in range(1, len(crosstab_result.columns) + 1)]

count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


# inside_append_dataframe_with_blank_rows(output_path, df_modified)
inside_append_dataframe_with_blank_rows(output_path, count1)














# Error15: Frequency of number of brands selected  by each interviewer in ever smoked in last 1 week(Q20)

Error_n = ['Error15: Frequency of number of brands selected  by each interviewer in ever smoked in last 1 week(Q20)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()
string="A_Q20_1,A_Q20_2,A_Q20_3,A_Q20_4,A_Q20_5,A_Q20_6,A_Q20_7,A_Q20_8,A_Q20_9,A_Q20_10,A_Q20_11,A_Q20_12,A_Q20_13,A_Q20_14,A_Q20_15,A_Q20_16,A_Q20_17,A_Q20_18,A_Q20_19,A_Q20_20,A_Q20_21,A_Q20_22,A_Q20_23,A_Q20_24,A_Q20_25,A_Q20_26,A_Q20_27,A_Q20_28,A_Q20_29,A_Q20_30,A_Q20_31,A_Q20_32,A_Q20_33,A_Q20_34,A_Q20_35,A_Q20_36,A_Q20_37,A_Q20_38,A_Q20_39,A_Q20_40,A_Q20_41,A_Q20_42,A_Q20_43,A_Q20_44,A_Q20_45,A_Q20_46,A_Q20_47,A_Q20_48,A_Q20_49,A_Q20_50,A_Q20_51,A_Q20_52,A_Q20_53,A_Q20_54,A_Q20_55,A_Q20_56,A_Q20_57,A_Q20_58,A_Q20_59,A_Q20_60,A_Q20_61,A_Q20_62,A_Q20_63,A_Q20_64,A_Q20_65,A_Q20_66,A_Q20_67,A_Q20_68,A_Q20_69,A_Q20_70,A_Q20_71,A_Q20_72,A_Q20_73,A_Q20_74,A_Q20_75,A_Q20_76,A_Q20_77,A_Q20_78,A_Q20_79,A_Q20_80,A_Q20_81,A_Q20_82,A_Q20_83,A_Q20_84,A_Q20_85,A_Q20_86,A_Q20_87,A_Q20_88,A_Q20_89,A_Q20_90,A_Q20_91,A_Q20_92,A_Q20_93,A_Q20_94,A_Q20_95,A_Q20_96,A_Q20_97,A_Q20_98,A_Q20_99,A_Q20_100,A_Q20_101"
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
df_global['sum_columns'] = df_global[input_string2].sum(axis=1)
df_modified = df_global.loc[:,['Respondent_id','Interviewer','sum_columns']]

crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['sum_columns'])



# dynamic_headers = [f'brands{i}' for i in range(1, len(crosstab_result.columns) + 1)]

count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


# inside_append_dataframe_with_blank_rows(output_path, df_modified)
inside_append_dataframe_with_blank_rows(output_path, count1)














# Error16: No. of sticksof regular brand (Q23)

Error_n = [' Error16: No. of sticksof regular brand (Q23)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()



df_modified = df_global.loc[:,['Q23','Interviewer']]
crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['Q23'])
count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


inside_append_dataframe_with_blank_rows(output_path, count1)












# Error17: Difference between avg sticks smoked/day & regular brand stick smoked/day

Error_n = [' Error17: Difference between avg sticks smoked/day & regular brand stick smoked/day']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()


df_global['Difference'] = df_global['Q11_2'] - df_global['Q23']
df_modified = df_global.loc[:,['Difference','Interviewer']]
crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['Difference'])
count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


inside_append_dataframe_with_blank_rows(output_path, count1)










# # Error18: Price fluctuation between Previous  & regular brand 

# Error_n = ['Error18: Price fluctuation between Previous  & regular brand ']
# df_name= pd.DataFrame(Error_n) 
# existing_df = pd.read_excel(output_path)
# startrow = existing_df.shape[0] + 4
# with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#     df_name.to_excel(writer, startrow=startrow, index=False, header=False)



# df_global=df.copy()
# control_col=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,63,64,65,66,67,68,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101]


# cig_price={ 1:[120],
#             2:[100],
#             3:[165],
#             4:[0],
#             5:[40],
#             6:[40],
#             7:[48],
#             8:[0],
#             9:[80],
#             10:[65],
#             11:[40],
#             12:[47],
#             13:[65],
#             14:[48],
#             15:[43],
#             16:[0],
#             17:[40],
#             18:[40],
#             19:[0],
#             20:[40],
#             21:[48],
#             22:[0],
#             23:[165],
#             24:[150],
#             25:[165],
#             26:[165],
#             27:[165],
#             28:[100],
#             29:[0],
#             30:[50],
#             31:[50],
#             32:[60],
#             33:[51],
#             34:[70],
#             35:[0],
#             36:[90],
#             37:[89],
#             38:[89],
#             39:[89],
#             40:[85],
#             41:[55],
#             42:[89],
#             43:[0],
#             44:[200],
#             45:[100],
#             46:[100],
#             47:[95],
#             48:[165],
#             49:[165],
#             50:[165],
#             51:[95],
#             52:[150],
#             53:[95],
#             54:[95],
#             55:[95],
#             56:[95],
#             57:[95],
#             58:[60],
#             59:[50],
#             60:[59],
#             61:[59],
#             62:[0],
#             63:[180],
#             64:[45],
#             65:[0],
#             66:[165],
#             67:[100],
#             68:[95],
#             69:[60],
#             70:[165],
#             71:[330],
#             72:[0],
#             73:[165],
#             74:[95],
#             75:[165],
#             76:[0],
#             77:[47],
#             78:[49],
#             79:[0],
#             80:[95],
#             81:[59],
#             82:[80],
#             83:[50],
#             84:[50],
#             85:[94],
#             86:[40],
#             87:[60],
#             88:[0],
#             89:[70],
#             90:[70],
#             91:[80],
#             92:[55],
#             93:[0],
#             94:[55],
#             95:[55],
#             96:[55],
#             97:[0],
#             98:[47],
#             99:[69],
#             100:[58],
#             101:[0],
#             102:[49],
#             103:[45],
#             104:[49],
#             105:[0],
#             106:[50],
#             107:[50],
#             108:[80],
#             109:[45],
#             110:[50],
#             111:[58],
#             112:[100],
#             113:[120],
#             114:[58],
#             115:[0],
#             116:[100],
#             117:[100],
#             118:[0],
#             119:[54],
#             120:[60],
#             121:[60],
#             122:[60],
#             123:[60],
#             124:[60],
#             125:[70],
#             126:[60],
#             127:[70],
#             128:[60],
#             129:[70],
#             130:[75],
#             131:[100],
#             132:[160],
#             133:[0],
#             134:[60],
#             135:[60],
#             136:[60],
#             137:[60],
#             138:[80],
#             139:[47],
#             140:[0],
#             141:[95],
#             142:[110],
#             143:[95],
#             144:[50],
#             145:[70],
#             146:[60],
#             147:[95],
#             148:[68],
#             149:[70],
#             150:[100]
#             }

# df_global = df_global.dropna(subset=['Q24'])

# df_global.reset_index(drop=True, inplace=True)


# for df_row in range (len(df_global)):
#     # if (cig_price[df_global.loc[df_row,'Q21']][1])-(cig_price[df_global.loc[df_row,'Q24']][1])<=40 and (cig_price[df_global.loc[df_row,'Q21']][1])-(cig_price[df_global.loc[df_row,'Q24']][1])>=-40:
#     # if (cig_price[df_global.loc[df_row,'Q21']][1])-(cig_price[df_global.loc[df_row,'Q24']][1])<=40  :
        
#     # print(cig_price[df_global.loc[df_row,'Q21']][0])
#     if (cig_price[df_global.loc[df_row,'Q21']][0])-(cig_price[df_global.loc[df_row,'Q24']][0])<=40 and (cig_price[df_global.loc[df_row,'Q21']][0])-(cig_price[df_global.loc[df_row,'Q24']][0])>=-40:
    
#         df_global.drop(df_row,inplace=True)
#         # print(df_global.loc[:,['Q21','Q24','Q27']])
#     else:
#         pass
    
    
# df_modified = df_global.loc[:,['Respondent_ID','Resp_Name','Interviewer','Q21','Q24']]


# input_string1=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,63,64,65,66,67,68,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101,102,103,104,105,106,107,108,109,110,111,112,113,114,115,116,117,118,119,120,121,122,123,124,125,126,127,128,129,130,131,132,133,134,135,136,137,138,139,140,141,142,143,144,145,146,147,148]
# input_string = "Amercian Club Length - 84 mm MRP/10's - Rs.120,American Club<br>Length - 84 mm<br>MRP/10's - Rs.100,Benson & Hedges <br>Length - 84 mm<br>MRP/10's - Rs.165,Blue Charms (Unspecified),Blue Charms 64<br>Length - 64 mm<br>MRP/10's - Rs.40,Blue Charms Regular 64<br>Length - 64 mm<br>MRP/10's - Rs.40,Blue Charms<br>Length - 64 mm<br>MRP/10's - Rs.48,Capstan (Unspecified),Capstan Excel<br>Length - 69 mm<br>MRP/10's - Rs.80,Capstan Filter<br>Length - 64 mm<br>MRP/10's - Rs.65,Capstan Metro<br>Length - 64 mm<br>MRP/10's - Rs.40,Capstan Pilot<br>Length - 64 mm<br>MRP/10's - Rs.47,Capstan Special Filter<br>Length - 64 mm<br>MRP/10's - Rs.65,Cavanders Gold Rich Taste<br>Length - 64 mm<br>MRP/10's - Rs.48,Cavenders Special Filter<br>Length - 64 mm<br>MRP/10's - Rs.43,Charminar (Unspecified),Charminar Filter 64<br>Length - 64 mm<br>MRP/10's - Rs.40,Charminar Plains 64<br>Length - 64 mm<br>MRP/10's - Rs.40,Charms (Unspecified),Charms Regular Filter<br>Length - 64 mm<br>MRP/10's - Rs.40,Charms Virginia Filter<br>Length - 64 mm<br>MRP/10's - Rs.48,Classic (Unspecified),Classic-20'S<br>Length - 84 mm<br>MRP/10's - Rs.165,Classic Connect<br>Length - 97 mm<br>MRP/10's - Rs.150,Classic Ice Burst<br>Length - 84 mm<br>MRP/10's - Rs.165,Classic Mild<br>Length - 84 mm<br>MRP/10's - Rs.165,Classic<br>Length - 84 mm<br>MRP/10's - Rs.165,Editions Trio<br>Length - 92 mm<br>MRP/10's - Rs.100,Flake (Unspecified),Flake Fine Cut<br>Length - 64 mm<br>MRP/10's - Rs.50,Flake Fresh<br>Length - 64 mm<br>MRP/10's - Rs.50,Flake Special Filter<br>Length - 64 mm<br>MRP/10's - Rs.60,Focus<br>Length - 64 mm<br>MRP/10's - Rs.51,Focus<br>Length - 69 mm<br>MRP/10's - Rs.70,Four Square (Unspecified),Four Square Clove Crush<br>Length - 64 mm<br>MRP/10's - Rs.90,Four Square Clove Crush<br>Length - 69 mm<br>MRP/10's - Rs.89,Four Square Crush Pan<br>Length - 69 mm<br>MRP/10's - Rs.89,Four Square Crush<br>Length - 69 mm<br>MRP/10's - Rs.89,Four Square Special<br>Length - 69 mm<br>MRP/10's - Rs.85,Four Square<br>Length - 69 mm<br>MRP/10's - Rs.55,Four Square<br>Length - 69 mm<br>MRP/10's - Rs.89,Gold Flake (Unspecified),Gold Flake Clove Filter<br>Length - 84 mm<br>MRP/10's - Rs.200,Gold Flake Filter<br>Length - 69 mm<br>MRP/10's - Rs.100,Gold Flake Indie Mint<br>Length - 69 mm<br>MRP/10's - Rs.100,Gold Flake Indie Mint<br>Length - 69 mm<br>MRP/10's - Rs.95,Gold Flake Kings Lights<br>Length - 84 mm<br>MRP/10's - Rs.165,Gold Flake Kings<br>Length - 84 mm<br>MRP/10's - Rs.165,Gold Flake Mix Pod<br>Length - 84 mm<br>MRP/10's - Rs.165,Gold Flake Neo Smart<br>Length - 69 mm<br>MRP/10's - Rs.95,Gold Flake Neo -Super Slim<br>Length - 97 mm<br>MRP/10's - Rs.150,Gold Flake Premium Filter Mint<br>Length - 69 mm<br>MRP/10's - Rs.95,Gold Flake Premium Filter<br>Length - 69 mm<br>MRP/10's - Rs.95,Gold Flake Premium Fresh Mint<br>Length - 69 mm<br>MRP/10's - Rs.95,Gold Flake Premium Mint Switch<br>Length - 69 mm<br>MRP/10's - Rs.95,Gold Flake Premium<br>Length - 69 mm<br>MRP/10's - Rs.95,Gold Flake Special<br>Length - 64 mm<br>MRP/10's - Rs.60,Gold Flake Super Star<br>Length - 64 mm<br>MRP/10's - Rs.50,Gold Flake Super Star<br>Length - 64 mm<br>MRP/10's - Rs.59,Gold Flake Superstar Mint<br>Length - 64 mm<br>MRP/10's - Rs.59,Gold Magic<br>Length - 84 mm,India King<br>Length - 84 mm<br>MRP/10's - Rs.180,Kingston 64<br>Length - 64 mm<br>MRP/10's - Rs.45,Marlboro (Unspecified),Marlboro <br>Length - 84 mm<br>MRP/10's - Rs.165,Marlboro Advance Compact<br>Length - 69 mm<br>MRP/10's - Rs.100,Marlboro Advance Compact<br>Length - 69 mm<br>MRP/10's - Rs.95,Marlboro Advance Pocket<br>Length - 64 mm<br>MRP/10's - Rs.60,Marlboro Advance<br>Length - 84 mm<br>MRP/10's - Rs.165,Marlboro Advance<br>Length - 84 mm<br>MRP/10's - Rs.330,Marlboro Clove<br>Length - 84 mm,Marlboro Clove<br>Length - 84 mm<br>MRP/10's - Rs.165,Marlboro Compact<br>Length - 69 mm<br>MRP/10's - Rs.95,Marlboro<br>Length - 84 mm<br>MRP/10's - Rs.165,Moments (Unspecified),Moments Regular 64<br>Length - 64 mm<br>MRP/10's - Rs.47,Moments Satins 64<br>Length - 64 mm<br>MRP/10's - Rs.49,Navy Cut (Unspecified),Navy Cut 69<br>Length - 69 mm<br>MRP/10's - Rs.95,Navy Cut Fruit<br>Length - 69 mm<br>MRP/10's - Rs.59,Navy Cut Rsft<br>Length - 69 mm<br>MRP/10's - Rs.80,Navy Cut Virginia Filter<br>Length - 64 mm<br>MRP/10's - Rs.50,Navy Cut Virginia/ Centure<br>Length - 64 mm<br>MRP/10's - Rs.50,Navy Cut<br>Length - 74 mm<br>MRP/10's - Rs.94,Originals<br>Length - 64 mm<br>MRP/10's - Rs.40,Pall Mall<br>Length - 64 mm<br>MRP/10's - Rs.60,Panama<br>Length - 64 mm,Players Fruity Cool / Players Minty Cool<br>Length - 69 mm<br>MRP/10's - Rs.70,Players<br>Length - 69 mm<br>MRP/10's - Rs.70,Players<br>Length - 69 mm<br>MRP/10's - Rs.80,R&W (M)<br>Length - 64 mm<br>MRP/10's - Rs.55,Red And White (Unspecified)<br>,Red And White Flake<br>Length - 69 mm<br>MRP/10's - Rs.55,Red And White Mint Burst<br>Length - 64 mm<br>MRP/10's - Rs.55,Red And White<br>Length - 69 mm<br>MRP/10's - Rs.55,Red Charms (Unspecified),Red Charms Mini Kings<br>Length - 64 mm<br>MRP/10's - Rs.47,Red Charms<br>Length - 69 mm<br>MRP/10's - Rs.69,Red Special Longs 69<br>Length - 69 mm<br>MRP/10's - Rs.58,Regent<br>Length - 84 mm,Royal<br>Length - 64 mm<br>MRP/10's - Rs.49,Select Flake Premium<br>Length - 64 mm<br>MRP/10's - Rs.45,Select<br>Length - 64 mm<br>MRP/10's - Rs.49,Silk Cut (Unspecified),Silk Cut Blue<br>Length - 64 mm<br>MRP/10's - Rs.50,Silk Cut Filter<br>Length - 64 mm<br>MRP/10's - Rs.50,Silk Cut Filter<br>Length - 69 mm<br>MRP/10's - Rs.80,Silk Cut Virginia<br>Length - 64 mm<br>MRP/10's - Rs.45,Silk Cut<br>Length - 64 mm<br>MRP/10's - Rs.50,Stellar <br>Length - 69 mm<br>MRP/10's - Rs.58,Stellar Define/Shift<br>Length - 99 mm<br>MRP/10's - Rs.100,Stellar Define<br>Length - 84 mm<br>MRP/10's - Rs.120,Stellar<br>Length - 69 mm<br>MRP/10's - Rs.58,Steller (Unspecified),Steller Define<br>Length - 84 mm<br>MRP/10's - Rs.100,Steller Shift<br>Length - 84 mm<br>MRP/10's - Rs.100,Total (Unspecified),Total Active Mint<br>Length - 64 mm<br>MRP/10's - Rs.54,Total Active Mint<br>Length - 64 mm<br>MRP/10's - Rs.60,Total Fruity Five<br>Length - 69 mm<br>MRP/10's - Rs.60,Total Fusion<br>Length - 69 mm<br>MRP/10's - Rs.60,Total Red White<br>Length - 69 mm<br>MRP/10's - Rs.60,Total Refresh<br>Length - 69 mm<br>MRP/10's - Rs.60,Total Refresh<br>Length - 69 mm<br>MRP/10's - Rs.70,Total Royal Twist<br>Length - 69 mm<br>MRP/10's - Rs.60,Total Royal Twist<br>Length - 69 mm<br>MRP/10's - Rs.70,Total Spearmint<br>Length - 69 mm<br>MRP/10's - Rs.60,Total Spearmint<br>Length - 69 mm<br>MRP/10's - Rs.70,Total Spearmint<br>Length - 69 mm<br>MRP/10's - Rs.75,Total T3<br>Length - 92 mm<br>MRP/10's - Rs.100,Total T3<br>Length - 92 mm<br>MRP/10's - Rs.160,Wave (Unspecified),Wave Cool Mint<br>Length - 69 mm<br>MRP/10's - Rs.60,Wave Fruity<br>Length - 69 mm<br>MRP/10's - Rs.60,Wave Mint<br>Length - 69 mm<br>MRP/10's - Rs.60,Wave<br>Length - 69 mm<br>MRP/10's - Rs.60,Wft-Reg<br>Length - 69 mm<br>MRP/10's - Rs.80,White Special Extra Smooth 64<br>Length - 64 mm<br>MRP/10's - Rs.47,Wills (Unspecified),Wills Filter<br>Length - 69 mm<br>MRP/10's - Rs.95,Wills Filter<br>Length - 74 mm<br>MRP/10's - Rs.110,Wills Flake Filter 69<br>Length - 69 mm<br>MRP/10's - Rs.95,Wills Flake Filter Gold Crest<br>Length - 64 mm<br>MRP/10's - Rs.50,Wills Flake Filter<br>Length - 64 mm<br>MRP/10's - Rs.70,Wills Flake Special Filter<br>Length - 69 mm<br>MRP/10's - Rs.60,Wills Navy Cut Delux<br>Length - 69 mm<br>MRP/10's - Rs.95,Zaffran 69<br>Length - 69 mm<br>MRP/10's - Rs.68"
# modified_string = input_string.split(",")

# dictionary = dict(zip(input_string1, modified_string))

# df_modified['Q21'] = df_modified['Q21'].replace(dictionary)
# df_modified['Q24'] = df_modified['Q24'].replace(dictionary)

    
# df_empty= pd.DataFrame()  
# df_s = df_modified['Interviewer'].value_counts()
    
# interviewererror=pd.DataFrame(df_s)    

# total_sum = interviewererror.sum()
# total_sum.name = 'Total'
# interviewererror = interviewererror._append(total_sum)
# interviewererror['Total']=interviewererror.sum(min_count=1,axis=1)
   



# df_name= pd.DataFrame(Error_n)  
# df_empty.to_csv(output_path,mode='a', index=True)
# df_empty.to_csv(output_path,mode='a', index=True)
# df_empty.to_csv(output_path,mode='a', index=True)
# df_empty.to_csv(output_path,mode='a', index=True)
# df_empty.to_csv(output_path,mode='a', index=True)
# df_empty.to_csv(output_path,mode='a', index=True)
# df_name.to_csv(output_path,mode='a', index=False, header=False)

# df_empty.to_csv(output_path,mode='a', index=True)
# df_empty.to_csv(output_path,mode='a', index=True)
# df_modified.to_csv(output_path,mode='a', index=False)
# df_empty.to_csv(output_path,mode='a', index=True)
# df_empty.to_csv(output_path,mode='a', index=True)
# interviewererror.to_csv(output_path, mode='a', index=True, )














# Error21: Percentage of loose vs pack

Error_n = ['Error21: Percentage of loose vs pack ']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)

df_global=df.copy()
input_string1=[1,2]
input_string2=['Loose','Packet']
dictionary = dict(zip(input_string1, input_string2))
df_global['Q27'] = df_global['Q27'].replace(dictionary)



df_percent = df_global['Q27'].value_counts(normalize=True) * 100  
df_percent = df_percent.round(0).astype(int)
Df_buyerspercen=pd.DataFrame(df_percent)   
inside_append_dataframe_with_blank_rows(output_path, Df_buyerspercen)





# Error22: Sticks, a pack consist of?

Error_n = ['Error22: Sticks, a pack consist of?']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)

df_global=df.copy()
df_filtered = df_global[df_global['Q29_1'] != -1]
df_modified = df_filtered.loc[:, ['Q29_1', 'Interviewer']]
crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['Q29_1'])
count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


inside_append_dataframe_with_blank_rows(output_path, count1)




# Error23: Mild or strong? (Q31)

Error_n = [' Error23: Mild or strong? (Q31)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)

df_global=df.copy()

df_global=df.copy()
input_string1=[1,2]
input_string2=['Strong','Mild']
dictionary = dict(zip(input_string1, input_string2))
df_global['Q31'] = df_global['Q31'].replace(dictionary)


df_modified = df_global.loc[:,['Q31','Interviewer']]
crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['Q31'])
count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total
inside_append_dataframe_with_blank_rows(output_path, count1)




df_percent = df_global['Q31'].value_counts(normalize=True) * 100  
df_percent = df_percent.round(0).astype(int)
Df_buyerspercen=pd.DataFrame(df_percent)   
inside_append_dataframe_with_blank_rows(output_path, Df_buyerspercen)









# Error24: Regualr brand is Mild or strong? (Q32)

Error_n = [' Error24: Regualr brand is Mild or strong? (Q32)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)

df_global=df.copy()

columns_to_process = [f'T_Q32_{i}' for i in range(1, 102)]

# Replace -1 with NaN in the specified columns
df_global[columns_to_process] = df_global[columns_to_process].replace(-1, np.nan)
df_global[columns_to_process] = df_global[columns_to_process].replace(1, 'Strong')
df_global[columns_to_process] = df_global[columns_to_process].replace(2, 'Mild')






# List of columns to process
columns_to_process = [f'T_Q32_{i}' for i in range(1, 102)]

# Initialize a dictionary to store the counts
counts = {'Brand': [], 'Strong': [], 'Mild': []}

# Loop through each column and count the occurrences of 1 and 2
for column in columns_to_process:
    strong_count = (df_global[column] =='Strong').sum()
    mild_count = (df_global[column] == 'Mild').sum()
    counts['Brand'].append(column)
    counts['Strong'].append(strong_count)
    counts['Mild'].append(mild_count)

# Create a summary DataFrame
summary_df = pd.DataFrame(counts)

string1="T_Q32_1,T_Q32_2,T_Q32_3,T_Q32_4,T_Q32_5,T_Q32_6,T_Q32_7,T_Q32_8,T_Q32_9,T_Q32_10,T_Q32_11,T_Q32_12,T_Q32_13,T_Q32_14,T_Q32_15,T_Q32_16,T_Q32_17,T_Q32_18,T_Q32_19,T_Q32_20,T_Q32_21,T_Q32_22,T_Q32_23,T_Q32_24,T_Q32_25,T_Q32_26,T_Q32_27,T_Q32_28,T_Q32_29,T_Q32_30,T_Q32_31,T_Q32_32,T_Q32_33,T_Q32_34,T_Q32_35,T_Q32_36,T_Q32_37,T_Q32_38,T_Q32_39,T_Q32_40,T_Q32_41,T_Q32_42,T_Q32_43,T_Q32_44,T_Q32_45,T_Q32_46,T_Q32_47,T_Q32_48,T_Q32_49,T_Q32_50,T_Q32_51,T_Q32_52,T_Q32_53,T_Q32_54,T_Q32_55,T_Q32_56,T_Q32_57,T_Q32_58,T_Q32_59,T_Q32_60,T_Q32_61,T_Q32_62,T_Q32_63,T_Q32_64,T_Q32_65,T_Q32_66,T_Q32_67,T_Q32_68,T_Q32_69,T_Q32_70,T_Q32_71,T_Q32_72,T_Q32_73,T_Q32_74,T_Q32_75,T_Q32_76,T_Q32_77,T_Q32_78,T_Q32_79,T_Q32_80,T_Q32_81,T_Q32_82,T_Q32_83,T_Q32_84,T_Q32_85,T_Q32_86,T_Q32_87,T_Q32_88,T_Q32_89,T_Q32_90,T_Q32_91,T_Q32_92,T_Q32_93,T_Q32_94,T_Q32_95,T_Q32_96,T_Q32_97,T_Q32_98,T_Q32_99,T_Q32_100,T_Q32_101"
input_string1 = string1.split(',')  
input_string1 = [item.strip() for item in input_string1]
string="American Club ,Cavanders Gold Rich Taste ,Chancellor ,Charminar Filter ,Charms Special Blue ,Classic ,Editions Trio ,Flake Special Filter ,Focus Mint  ,Gold Flake Kings ,Gold Flake Premium ,Gold Flake Premium Neo Smart ,Gold Flake Indie Mint ,Gold Flake Special ,Gold Flake Special Mint ,Gold Flake Super Star ,India King ,Marlboro KSFT ,Marlboro Advance Compact ,Marlboro Pocket Filter ,Navy Cut Fruit ,Navy Cut Virginia Filter ,Red & White Select ,Red & White Select,Regent ,Regent Black ,Regent Cool ,Silk Cut Blue ,Silk Cut Filter ,Silk Cut Virginia ,Special Blues ,Special Red Longs ,Special Red Signature ,Stellar Cool Blast  ,T3 White  ,Total Refresh ,Total Royal Twist ,Total Spearmint ,Wave Cool Mint ,Wills Navy Cut Filter ,Wills Navy Cut,American Club Mint ,American Club Smash ,American Fruit ,Benson & Hedges  ,Berkeley ,Berkely ,Blue Charms ,Blue Charms ,Bristol ,Cavander Gold  ,Chancellor ,Charminar Plains ,Charms Regular Filter ,Classic Connect ,Classic Ice Burst ,Club One ,Duke ,Editions Active Mint ,Editions Ice Fruit ,Editions Spark ,Flake Excel ,Flake Liberty ,Flake Mint Switch ,Flake Nova ,Flake White ,Gold Flake Century ,Gold Flake Filter ,Gold Flake Filter ,Gold Flake Kings ,Gold Flake Kings Lights ,Gold Flake Kings SLK ,Golden Gold Flake ,Kingston ,Marlboro Clove ,Marlboro Fuse Beyond ,NATIONAL GOLD FLAKE ,Navy Cut Deluxe Filter ,Originals ,Panama ,Panama Filter ,Panama Filter ,Player's Fruit ,Player's Mint ,Royal ,Scissors Menthol  ,Classic ,Stellar Slims Define ,Stellar Slims Shift ,Total Active Mint ,Wave Fruity ,Wave Mint ,Will Flake Premium Filter ,Zaffran ,American Club Clove Magik ,Classic AlphaTec ,Classic Double Burst ,Classic Verve ,Stellar Define Pan ,Stellar Shift Duos ,Wills Insignia "
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
dictionary = dict(zip(input_string1, input_string2))

summary_df['Brand'] = summary_df['Brand'].replace(dictionary)

inside_append_dataframe_with_blank_rows(output_path, summary_df)









# # Error25: Number of brands being coded across all attributes(B1)

# Error_n = [' Error25: Number of brands being coded across all attributes(B1)']
# df_name= pd.DataFrame(Error_n) 
# existing_df = pd.read_excel(output_path)
# startrow = existing_df.shape[0] + 4
# with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#     df_name.to_excel(writer, startrow=startrow, index=False, header=False)

# df_global=df.copy()





# Error26: Number of times Statements being coded (C1)

Error_n = [' Error26: Number of times Statements being coded (C1)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)

df_global=df.copy()

# result = df_global.groupby('Interviewer')[input_string1].sum()







string1="A_C1_1,A_C1_2,A_C1_3,A_C1_4,A_C1_5,A_C1_6,A_C1_7,A_C1_8,A_C1_9,A_C1_10,A_C1_11,A_C1_12,A_C1_13,A_C1_14,A_C1_15,A_C1_16"
input_string1 = string1.split(',')  
input_string1 = [item.strip() for item in input_string1]
string="It offers satisfying smoking experience,It has good taste,It has good After Taste – taste that remains in mouth after smoking,Smooth smoking experience,It has right amount of harshness,Good strength / Just the right strength,Good hit / kick / impact,Mouthful of smoke / Volume of smoke which comes in the mouth when you take a puff,It has right amount of effort required to inhale smoke,Cigarette looks good/ stylish,Cigarette pack looks good/ stylish,Good quality of filter,Length of cigarette is right / good,Keeps smell of my breath pleasant even after smoking,Like different capsule flavour of cigarette,Cooling sensation in throat of the capsule "
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
dictionary = dict(zip(input_string1, input_string2))

# df_global = df_global.rename(columns=dictionary, inplace=True)


result = df_global.groupby('Interviewer')[input_string1].sum()
result = result.rename(columns=dictionary)

result['Total']=result.sum(min_count=1,axis=1)
total=result.sum()
total.name='Total'
result=result._append(total)
inside_append_dataframe_with_blank_rows(output_path, result)











# Error27: Statement selection (C2)

Error_n = ['  Error27: Statement selection (C2)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)

df_global=df.copy()


# input_string1=[1,2,3,4,5]
# input_string2=['Ex Diss','Diss','N S/D', 'SS','ES']
# dictionary = dict(zip(input_string1, input_string2))
# df_global['Q31'] = df_global['Q31'].replace(dictionary)







columns_to_process = [f'T_C2_{i}' for i in range(1,6)]

# Replace -1 with NaN in the specified columns
df_global[columns_to_process] = df_global[columns_to_process].replace(-1, np.nan)
df_global[columns_to_process] = df_global[columns_to_process].replace(1, 'Ex Diss')
df_global[columns_to_process] = df_global[columns_to_process].replace(2, 'Diss')
df_global[columns_to_process] = df_global[columns_to_process].replace(3, 'N S/D')
df_global[columns_to_process] = df_global[columns_to_process].replace(4, 'SS')
df_global[columns_to_process] = df_global[columns_to_process].replace(5, 'ES')








# List of columns to process
columns_to_process = [f'T_C2_{i}' for i in range(1, 6)]

# Initialize a dictionary to store the counts
counts = {'Statement': [], 'Ex Diss': [], 'Diss': [],'N S/D': [],'SS': [],'ES': []}

# Loop through each column and count the occurrences of 1 and 2
for column in columns_to_process:
    Ex_Diss_count = (df_global[column] =='Ex Diss').sum()
    Diss_count = (df_global[column] == 'Diss').sum()
    NSD_count = (df_global[column] =='N S/D').sum()
    
    SS_count = (df_global[column] =='SS').sum()
    
    ES_count = (df_global[column] =='ES').sum()
   
    counts['Statement'].append(column)
    counts['Ex Diss'].append(Ex_Diss_count)
    counts['Diss'].append(Diss_count)
    counts['N S/D'].append(NSD_count)
    counts['SS'].append(SS_count)
    counts['ES'].append(ES_count)

# Create a summary DataFrame
summary_df = pd.DataFrame(counts)

string1="T_C2_1,T_C2_2,T_C2_3,T_C2_4,T_C2_5,T_C2_3,T_C2_4,T_C2_5,T_C2_4,T_C2_5"
input_string1 = string1.split(',')  
input_string1 = [item.strip() for item in input_string1]
string="Overall satisfaction  ,Quality of filter ,Taste of the cigarette ,After Taste ,Flavour of Capsule "
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
dictionary = dict(zip(input_string1, input_string2))

summary_df['Statement'] = summary_df['Statement'].replace(dictionary)

inside_append_dataframe_with_blank_rows(output_path, summary_df)























# Error28: Just Right scale (C3)

Error_n = [' Error28: Just Right scale (C3)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)

df_global=df.copy()


# input_string1=[1,2,3,4,5]
# input_string2=['Ex Diss','Diss','N S/D', 'SS','ES']
# dictionary = dict(zip(input_string1, input_string2))
# df_global['Q31'] = df_global['Q31'].replace(dictionary)







columns_to_process = [f'T_C3_{i}' for i in range(1,8)]

# Replace -1 with NaN in the specified columns
df_global[columns_to_process] = df_global[columns_to_process].replace(-1, np.nan)
df_global[columns_to_process] = df_global[columns_to_process].replace(1, 'Too Less')
df_global[columns_to_process] = df_global[columns_to_process].replace(2, 'Slightly less')
df_global[columns_to_process] = df_global[columns_to_process].replace(3, 'JR')
df_global[columns_to_process] = df_global[columns_to_process].replace(4, 'Slightly More')
df_global[columns_to_process] = df_global[columns_to_process].replace(5, 'Too Much')








# List of columns to process
columns_to_process = [f'T_C3_{i}' for i in range(1,8)]

# Initialize a dictionary to store the counts
counts = {'Statement': [], 'Too Less': [], 'Slightly less': [],'JR': [],'Slightly More': [],'Too Much': []}

# Loop through each column and count the occurrences of 1 and 2
for column in columns_to_process:
    To_less_count = (df_global[column] =='Too Less').sum()
    Slightless_count = (df_global[column] == 'Slightly less').sum()
    JR_count = (df_global[column] =='JR').sum()
    
    Slightmore_count = (df_global[column] =='Slightly More').sum()
    
    Toomuch_count = (df_global[column] =='Too Much').sum()
   
    counts['Statement'].append(column)
    counts['Too Less'].append(To_less_count)
    counts['Slightly less'].append(Slightless_count)
    counts['JR'].append(JR_count)
    counts['Slightly More'].append(Slightmore_count)
    counts['Too Much'].append(Toomuch_count)

# Create a summary DataFrame
summary_df = pd.DataFrame(counts)

string1="T_C3_1,T_C3_2,T_C3_3,T_C3_4,T_C3_5,T_C3_6,T_C3_7"
input_string1 = string1.split(',')  
input_string1 = [item.strip() for item in input_string1]
string="Mouthful of smoke ,Effort required to inhale smoke ,Hit or kick,Smooth smoking experience ,Harshness of the cigarette ,Strength of the cigarette ,Length of the cigarette  "
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
dictionary = dict(zip(input_string1, input_string2))

summary_df['Statement'] = summary_df['Statement'].replace(dictionary)

inside_append_dataframe_with_blank_rows(output_path, summary_df)


















# Error30: Reasons to smoke regular brand (D1)

Error_n = [' Error30: Reasons to smoke regular brand (D1)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)

df_global=df.copy()



string1="A_D1_2,A_D1_3,A_D1_5,A_D1_6,A_D1_8,A_D1_9,A_D1_10,A_D1_11,A_D1_12,A_D1_13,A_D1_15,A_D1_16,A_D1_18,A_D1_19,A_D1_20,A_D1_21,A_D1_22,A_D1_24,A_D1_25,A_D1_26,A_D1_28,A_D1_29,A_D1_30,A_D1_32,A_D1_33"
input_string1 = string1.split(',')  
input_string1 = [item.strip() for item in input_string1]
string="I liked this cigarette more,I have been smoking this cigarette since a long time,High quality cigarette,Everything taken together the brand is appealing to me,It offers satisfying smoking experience,Good taste / Better taste,Smooth smoking experience,Good strength/ Just the right strength,Good hit/ kick/ impact,Good Mouthful of smoke / Volume of smoke which comes in the mouth when you take a puff,Cigarette looks good/ stylish,Cigarette pack looks good/ stylish,Good quality of filter,Length is right / good,Keeps smell of my breath pleasant even after smoking,Can change flavour with capsules,Cooling sensation in throat,Price is affordable/ within my budget,It is for people like me,It is a better brand of cigarette,Friends / Colleagues recommend this brand / smoke this brand,Shopkeepers recommended this brand,It is a very popular cigarette in the market,This cigarette is easily available in the market,This cigarette is available in the shop near my home/ office/ others "
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
dictionary = dict(zip(input_string1, input_string2))

# df_global = df_global.rename(columns=dictionary, inplace=True)


result = df_global.groupby('Interviewer')[input_string1].sum()
result = result.rename(columns=dictionary)

result['Total']=result.sum(min_count=1,axis=1)
total=result.sum()
total.name='Total'
result=result._append(total)
inside_append_dataframe_with_blank_rows(output_path, result)









# Error: Awareness of Cig with Capsule (H1)

Error_n = [' Error: Awareness of Cig with Capsule (H1)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)

df_global=df.copy()

input_string1=[1,2]

input_string2 = ['Yes','No']
dictionary = dict(zip(input_string1, input_string2))
df_global['H1'] = df_global['H1'].replace(dictionary)
df_global['H1'] = df_global['H1'].replace(-1, np.nan)


df_modified = df_global.loc[:,['H1','Interviewer']]
crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['H1'])
count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


inside_append_dataframe_with_blank_rows(output_path, count1)













# Error: Capsule bursting (H5)

Error_n = [' Error: Capsule bursting (H5)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)

df_global=df.copy()

input_string1=[1,2,3,4]

input_string2 = ['Yes, always','Yes,mostly','No, mostly smoke without bursting ','Never']
dictionary = dict(zip(input_string1, input_string2))
df_global['H5'] = df_global['H5'].replace(dictionary)
df_global['H5'] = df_global['H5'].replace(-1, np.nan)


df_modified = df_global.loc[:,['H5','Interviewer']]
crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['H5'])
count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


inside_append_dataframe_with_blank_rows(output_path, count1)






# Error: Reasons not regularly smoke a Capsule cigarette (H7)

Error_n = [' Error: Reasons not regularly smoke a Capsule cigarette (H7)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)

df_global=df.copy()



string1="A_H7_2,A_H7_3,A_H7_5,A_H7_6,A_H7_8,A_H7_10,A_H7_11,A_H7_13,A_H7_14,A_H7_15,A_H7_16,A_H7_17,A_H7_18,A_H7_19,A_H7_21,A_H7_22,A_H7_24,A_H7_25,A_H7_26,A_H7_27,A_H7_28,A_H7_29,A_H7_31,A_H7_32,A_H7_34,A_H7_35,A_H7_36,A_H7_38,A_H7_39,A_H7_40"
input_string1 = string1.split(',')  
input_string1 = [item.strip() for item in input_string1]
string="Like my regular cigarette more,Smoking my regular cigarette since a long time,Overall capsule cigarette is not a good quality product,Capsule cigarette is not appealing to me,Smoked capsule cigarette in the past and it was not good,Don't know much about capsule cigarettes,Never tried capsule cigarette,Like taste of original tobacco,Does not offer satisfying smoking experience,Does not have good taste,Does not have smooth smoking experience,Does not have good strength/ does not have right strength,Does not provide right hit/ kick/ impact,Does not provide a good mouthful of smoke / Volume of smoke which comes in the mouth when you take a puff,Stick does not look good/ stylish,Pack does not look good/ stylish,No difference as compared to regular cigarettes,Does not have good quality of filter,Length of Capsule cigarette is small,Smell of the breath becomes unpleasant after smoking ,Capsule cigarette has limited flavours,It is difficult to burst capsules,Capsule cigarette is not value for money,Capsule cigarette is not for people like me,Friends/Colleagues do not recommend ,Shopkeepers do not recommend ,Not a very popular cigarette in the market,Capsule cigarette is not easily available in the market,Capsule cigarette is not available in the shop near my home/ office/ others,Others "
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
dictionary = dict(zip(input_string1, input_string2))

# df_global = df_global.rename(columns=dictionary, inplace=True)

for column in input_string1:

# Replace -1 with NaN in the specified columns
    df_global[column] = df_global[column].replace(-1, np.nan)

result = df_global.groupby('Interviewer')[input_string1].sum()
result = result.rename(columns=dictionary)

result['Total']=result.sum(min_count=1,axis=1)
total=result.sum()
total.name='Total'
result=result._append(total)
inside_append_dataframe_with_blank_rows(output_path, result)











# Error: Seen this PACKET of cigarette brand Blue Charms? (S3)

Error_n = [' Error: Seen this PACKET of cigarette brand Blue Charms? (S3)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)

df_global=df.copy()

input_string1=[1,2]

input_string2 = ['Yes','No/Not sure']
dictionary = dict(zip(input_string1, input_string2))
df_global['S3'] = df_global['S3'].replace(dictionary)
df_global['S3'] = df_global['S3'].replace(-1, np.nan)


df_modified = df_global.loc[:,['S3','Interviewer']]
crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['S3'])
count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


inside_append_dataframe_with_blank_rows(output_path, count1)













# Error: How much do you like the PACKET of this cigarette brand? (S4)

Error_n = [' Error: How much do you like the PACKET of this cigarette brand? (S4)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)

df_global=df.copy()

input_string1=[1,2,3,4,5]

input_string2 = ['Do not like it at all','Dislike it somewhat','Neither like it nor dislike it','Like it somewhat','Like it very much']
dictionary = dict(zip(input_string1, input_string2))
df_global['S4'] = df_global['S4'].replace(dictionary)
df_global['S4'] = df_global['S4'].replace(-1, np.nan)


df_modified = df_global.loc[:,['S4','Interviewer']]
crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['S4'])
count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


inside_append_dataframe_with_blank_rows(output_path, count1)












# Error: how much you agree or disagree with each of the statement (S5)

Error_n = ['  Error: how much you agree or disagree with each of the statement (S5)']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)

df_global=df.copy()


# input_string1=[1,2,3,4,5]
# input_string2=['Ex Diss','Diss','N S/D', 'SS','ES']
# dictionary = dict(zip(input_string1, input_string2))
# df_global['Q31'] = df_global['Q31'].replace(dictionary)







columns_to_process = [f'T_S5_{i}' for i in range(1,5)]

# Replace -1 with NaN in the specified columns
df_global[columns_to_process] = df_global[columns_to_process].replace(-1, np.nan)
df_global[columns_to_process] = df_global[columns_to_process].replace(1, 'Ex Diss')
df_global[columns_to_process] = df_global[columns_to_process].replace(2, 'Diss')
df_global[columns_to_process] = df_global[columns_to_process].replace(3, 'N S/D')
df_global[columns_to_process] = df_global[columns_to_process].replace(4, 'SS')
df_global[columns_to_process] = df_global[columns_to_process].replace(5, 'ES')








# List of columns to process
columns_to_process = [f'T_C2_{i}' for i in range(1, 6)]

# Initialize a dictionary to store the counts
counts = {'Statement': [], 'Ex Diss': [], 'Diss': [],'N S/D': [],'SS': [],'ES': []}

# Loop through each column and count the occurrences of 1 and 2
for column in columns_to_process:
    Ex_Diss_count = (df_global[column] =='Ex Diss').sum()
    Diss_count = (df_global[column] == 'Diss').sum()
    NSD_count = (df_global[column] =='N S/D').sum()
    
    SS_count = (df_global[column] =='SS').sum()
    
    ES_count = (df_global[column] =='ES').sum()
   
    counts['Statement'].append(column)
    counts['Ex Diss'].append(Ex_Diss_count)
    counts['Diss'].append(Diss_count)
    counts['N S/D'].append(NSD_count)
    counts['SS'].append(SS_count)
    counts['ES'].append(ES_count)

# Create a summary DataFrame
summary_df = pd.DataFrame(counts)

string1="T_S5_1,T_S5_2,T_S5_3,T_S5_4"
input_string1 = string1.split(',')  
input_string1 = [item.strip() for item in input_string1]
string="It is an Attractive pack,It is a Unique pack,Is a premium pack,It is an exciting pack "
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
dictionary = dict(zip(input_string1, input_string2))

summary_df['Statement'] = summary_df['Statement'].replace(dictionary)

inside_append_dataframe_with_blank_rows(output_path, summary_df)














# Error: GPS data absent

Error_n = ['  Error: GPS data absent']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)

df_global=df.copy()


missing_values = df.groupby('Interviewer')['Latitude'].apply(lambda x: x.isnull().sum())

inside_append_dataframe_with_blank_rows(output_path, missing_values)













# # Error: Click a Picture of the Respondent (P5)

# Error_n = [' Click a Picture of the Respondent (P5)']
# df_name= pd.DataFrame(Error_n) 
# existing_df = pd.read_excel(output_path)
# startrow = existing_df.shape[0] + 4
# with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#     df_name.to_excel(writer, startrow=startrow, index=False, header=False)

# df_global=df.copy()

# input_string1=[1,2]

# input_string2 = ['Open Camera','Refused']
# dictionary = dict(zip(input_string1, input_string2))
# df_global['S3'] = df_global['S3'].replace(dictionary)
# df_global['S3'] = df_global['S3'].replace(-1, np.nan)


# df_modified = df_global.loc[:,['S3','Interviewer']]
# crosstab_result = pd.crosstab(df_modified['Interviewer'], df_modified['S3'])
# count1=pd.DataFrame(crosstab_result )
# count1['Total']=count1.sum(min_count=1,axis=1)


# total=count1.sum()
# # total.name='Total'
# # count1=count1._append(total)
# count1.loc['Total']=total


# inside_append_dataframe_with_blank_rows(output_path, count1)














