from openpyxl import Workbook
from openpyxl import load_workbook
import pandas as pd
import numpy as np
from openpyxl.styles import Alignment

input_file = r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Data input\25.BHS.xlsx'
input_sheet = 'Sheet1'
output_path = r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Python output\Analysis\21Vizag_BHSanalysis.xlsx'

df = pd.read_excel(input_file,input_sheet)

entries_to_exclude = ['test', 'trr','demo','test']
df = df[~df['Interviewer'].isin(entries_to_exclude)]


ID_to_exclude = [210777004,210844981,211205743,210845893,211167584,211167602,211168044,211168049,211198899,211144726,211231256,211249825,211251795,211253025,210838938,210839088,210841644,210843453,210844810,210846789,210850696,210904706,210906670,210907729,210908512,210909983,210911099]
df = df[~df['SbjNum'].isin(ID_to_exclude)]



df=df[df['Q2'] == 4]
# df=df[df['Q1_1']==1]

# df=df[df['A_Segment_3']==1]








# Writing Error index in dataframe

# Error_n = ['Count1: Zone Quota ']
# df_name= pd.DataFrame(Error_n) 
# df_name.to_excel(output_path, index=False, header=False)


# Step 1: Create a list of topic names
topics = ['Table1 :Combined table of Counts for total awareness and Total usage','Table2:Shifting from previous to current regular brand','Table 3 : Count:Age wise brands','Table4 : Count:SEC wise brands','Table5 :Count:Price quoted for MOUB']


df_name = pd.DataFrame(topics, columns=['Logic list'])
df_name.to_excel(output_path, index=False)




def inside_append_dataframe_with_blank_rows(file_path, dataframe, blank_rows=2):
    """
    Appends a DataFrame to an existing Excel file with a specified number of blank rows in between.

    Parameters:
    - file_path: str, path to the Excel file
    - dataframe: pd.DataFrame, DataFrame to append
    - blank_rows: int, number of blank rows between appended DataFrames (default is 5)
    """
    
    with pd.ExcelWriter(
            file_path,
            engine='openpyxl',
            mode='a',
            if_sheet_exists='overlay') as writer:
        reader = pd.read_excel(file_path)
        dataframe.to_excel(
            writer,
            startrow=reader.shape[0] + blank_rows,
            index=True,
            header=True)

   
        
    workbook = load_workbook(file_path)
    worksheet = workbook.active

    # Center-align the text in the appended DataFrame
    for row in worksheet.iter_rows(min_row=startrow + 1, max_row=startrow + len(dataframe) + 5, min_col=1, max_col=len(dataframe.columns) + 1):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Save the workbook
    workbook.save(file_path)     







def outside_append_dataframe_with_blank_rows(file_path, dataframe, blank_rows=5):
    """
    Appends a DataFrame to an existing Excel file with a specified number of blank rows in between.

    Parameters:
    - file_path: str, path to the Excel file
    - dataframe: pd.DataFrame, DataFrame to append
    - blank_rows: int, number of blank rows between appended DataFrames (default is 5)
    """
    
    with pd.ExcelWriter(
            file_path,
            engine='openpyxl',
            mode='a',
            if_sheet_exists='overlay') as writer:
        reader = pd.read_excel(file_path)
        dataframe.to_excel(
            writer,
            startrow=reader.shape[0] + blank_rows,
            index=True,
            header=True)











# Count1: Tom  


# Error_n = ['Count1 :Tom']
# df_name= pd.DataFrame(Error_n) 
# existing_df = pd.read_excel(output_path)
# startrow = existing_df.shape[0] + 4
# with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#     df_name.to_excel(writer, startrow=startrow, index=False, header=False)




df_global=df.copy()
df_global.reset_index(drop=True, inplace=True)


df_dictformation = pd.read_excel(r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Data input\BHS 22-Oct-24_DataMap.xlsx','MOUB')



# Join two columns into a dictionary
# Assuming 'Column1' and 'Column2' are the names of the columns you want to join
dict_format = dict(zip(df_dictformation['Column'], df_dictformation['Brand']))





df_s = df_global['Q14'].value_counts()


Valuecount1=pd.DataFrame(df_s)
# total_sum = Valuecount1.sum()
# total_sum.name = 'Total_sum'
# Valuecount1=Valuecount1._append(total_sum)




Valuecount1.index = Valuecount1.index.to_series().replace(dict_format)








# Count 2 = Spont brand count






df_global=df.copy()

df_global.reset_index(drop=True, inplace=True)



df_dictformation = pd.read_excel(r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Data input\BHS 22-Oct-24_DataMap.xlsx','BrandcolumnQ14')



# Join two columns into a dictionary
# Assuming 'Column1' and 'Column2' are the names of the columns you want to join
dict_format = dict(zip(df_dictformation['Column'], df_dictformation['Brand']))
df_global.rename(columns=dict_format, inplace=True)

renamed_columns = list(dict_format.values())

# Adding a new column 'Row_Sum' that contains the sum of each row across the renamed columns
sum_of_columns= df_global[renamed_columns].sum()




Valuecount2=pd.DataFrame(sum_of_columns)







# Count 3 = Unaided awareness









Valuecount3 = pd.concat([Valuecount1, Valuecount2], axis=1)
Valuecount3.columns = ['TOM', 'Spont']
Valuecount3['Unaided awareness']=Valuecount3.sum(min_count=1,axis=1)






# Count 4 = Aided Awareness brand count






df_global=df.copy()

df_global.reset_index(drop=True, inplace=True)


df_dictformation = pd.read_excel(r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Data input\BHS 22-Oct-24_DataMap.xlsx','Brand15')



# Join two columns into a dictionary
# Assuming 'Column1' and 'Column2' are the names of the columns you want to join
dict_format = dict(zip(df_dictformation['Column'], df_dictformation['Brand']))

df_global.rename(columns=dict_format, inplace=True)

renamed_columns = list(dict_format.values())

# Adding a new column 'Row_Sum' that contains the sum of each row across the renamed columns
sum_of_columns= df_global[renamed_columns].sum()




Valuecount4=pd.DataFrame(sum_of_columns)






# Count 5 = Total Awareness





df_global=df.copy()
df_global.reset_index(drop=True, inplace=True)



df_dictformation = pd.read_excel(r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Data input\BHS 22-Oct-24_DataMap.xlsx','Brandtotalaware')



# Join two columns into a dictionary
# Assuming 'Column1' and 'Column2' are the names of the columns you want to join
dict_format = dict(zip(df_dictformation['Column'], df_dictformation['Brand']))
df_global.rename(columns=dict_format, inplace=True)

renamed_columns = list(dict_format.values())

# Adding a new column 'Row_Sum' that contains the sum of each row across the renamed columns
sum_of_columns= df_global[renamed_columns].sum()




Valuecount5=pd.DataFrame(sum_of_columns)








# Count 6 = Ever smoked brand count







df_global=df.copy()
df_global.reset_index(drop=True, inplace=True)



df_dictformation = pd.read_excel(r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Data input\BHS 22-Oct-24_DataMap.xlsx','Brandeversmoke(Q16)')



# Join two columns into a dictionary
# Assuming 'Column1' and 'Column2' are the names of the columns you want to join
dict_format = dict(zip(df_dictformation['Column'], df_dictformation['Brand']))
df_global.rename(columns=dict_format, inplace=True)

renamed_columns = list(dict_format.values())

# Adding a new column 'Row_Sum' that contains the sum of each row across the renamed columns
sum_of_columns= df_global[renamed_columns].sum()




Valuecount6=pd.DataFrame(sum_of_columns)












# Count 7 = Last 2 year smoked brand count



df_global=df.copy()
df_global.reset_index(drop=True, inplace=True)



df_dictformation = pd.read_excel(r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Data input\BHS 22-Oct-24_DataMap.xlsx','Brand(Q17)')



# Join two columns into a dictionary
# Assuming 'Column1' and 'Column2' are the names of the columns you want to join
dict_format = dict(zip(df_dictformation['Column'], df_dictformation['Brand']))
df_global.rename(columns=dict_format, inplace=True)

renamed_columns = list(dict_format.values())

# Adding a new column 'Row_Sum' that contains the sum of each row across the renamed columns
sum_of_columns= df_global[renamed_columns].sum()




Valuecount7=pd.DataFrame(sum_of_columns)





# Count 8 = Last 1 year smoked brand count



df_global=df.copy()
df_global.reset_index(drop=True, inplace=True)



df_dictformation = pd.read_excel(r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Data input\BHS 22-Oct-24_DataMap.xlsx','Q18')



# Join two columns into a dictionary
# Assuming 'Column1' and 'Column2' are the names of the columns you want to join
dict_format = dict(zip(df_dictformation['Column'], df_dictformation['Brand']))
df_global.rename(columns=dict_format, inplace=True)

renamed_columns = list(dict_format.values())

# Adding a new column 'Row_Sum' that contains the sum of each row across the renamed columns
sum_of_columns= df_global[renamed_columns].sum()




Valuecount8=pd.DataFrame(sum_of_columns)















# Count 9 = Last 1 month smoked brand count



df_global=df.copy()
df_global.reset_index(drop=True, inplace=True)



df_dictformation = pd.read_excel(r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Data input\BHS 22-Oct-24_DataMap.xlsx','Q19')



# Join two columns into a dictionary
# Assuming 'Column1' and 'Column2' are the names of the columns you want to join
dict_format = dict(zip(df_dictformation['Column'], df_dictformation['Brand']))
df_global.rename(columns=dict_format, inplace=True)

renamed_columns = list(dict_format.values())

# Adding a new column 'Row_Sum' that contains the sum of each row across the renamed columns
sum_of_columns= df_global[renamed_columns].sum()




Valuecount9=pd.DataFrame(sum_of_columns)










# Count 10 = Last 2 week smoked brand count



df_global=df.copy()
df_global.reset_index(drop=True, inplace=True)



df_dictformation = pd.read_excel(r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Data input\BHS 22-Oct-24_DataMap.xlsx','Q20.1')



# Join two columns into a dictionary
# Assuming 'Column1' and 'Column2' are the names of the columns you want to join
dict_format = dict(zip(df_dictformation['Column'], df_dictformation['Brand']))
df_global.rename(columns=dict_format, inplace=True)

renamed_columns = list(dict_format.values())

# Adding a new column 'Row_Sum' that contains the sum of each row across the renamed columns
sum_of_columns= df_global[renamed_columns].sum()




Valuecount10=pd.DataFrame(sum_of_columns)











# Count 11 = Last 1 week smoked brand count



df_global=df.copy()
df_global.reset_index(drop=True, inplace=True)



df_dictformation = pd.read_excel(r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Data input\BHS 22-Oct-24_DataMap.xlsx','Q20')



# Join two columns into a dictionary
# Assuming 'Column1' and 'Column2' are the names of the columns you want to join
dict_format = dict(zip(df_dictformation['Column'], df_dictformation['Brand']))
df_global.rename(columns=dict_format, inplace=True)

renamed_columns = list(dict_format.values())

# Adding a new column 'Row_Sum' that contains the sum of each row across the renamed columns
sum_of_columns= df_global[renamed_columns].sum()




Valuecount11=pd.DataFrame(sum_of_columns)








# Count 12 = MOUB




df_global=df.copy()
df_global.reset_index(drop=True, inplace=True)


df_dictformation = pd.read_excel(r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Data input\BHS 22-Oct-24_DataMap.xlsx','MOUB')



# Join two columns into a dictionary
# Assuming 'Column1' and 'Column2' are the names of the columns you want to join
dict_format = dict(zip(df_dictformation['Column'], df_dictformation['Brand']))





df_s = df_global['Q21'].value_counts()


Valuecount12=pd.DataFrame(df_s)
# total_sum = Valuecount1.sum()
# total_sum.name = 'Total_sum'
# Valuecount1=Valuecount1._append(total_sum)




Valuecount12.index = Valuecount12.index.to_series().replace(dict_format)






# Concat


Error_n = [' Table1 :Combined table of Counts for total awareness and Total usage']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)




resultant = pd.concat([Valuecount3,Valuecount4,Valuecount5,Valuecount6,Valuecount7,Valuecount8,Valuecount9,Valuecount10,Valuecount11,Valuecount12], axis=1)
resultant.columns = ['TOM', 'Spont', 'Unaided','Aided awareness','Total Awareness','Ever smoked','Last 2 year smoked','Last 1 year smoked', 'Last 1 month smoked', 'Last 2 weeks smoked','Last 1 weeks smoked','Regularly smoking' ]

total_sum = resultant.sum()
total_sum.name = 'Total_sum'
resultant=resultant._append(total_sum)





inside_append_dataframe_with_blank_rows(output_path, resultant)





















# Count = Shifting from previous to current regular brand 

Error_n = ['Table2:Shifting from previous to current regular brand  ']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)





df_dictformation = pd.read_excel(r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Data input\BHS 22-Oct-24_DataMap.xlsx','MOUB')



# Join two columns into a dictionary
# Assuming 'Column1' and 'Column2' are the names of the columns you want to join
dict_format = dict(zip(df_dictformation['Column'], df_dictformation['Brand']))




df_global=df.copy()
df_global.reset_index(drop=True, inplace=True)

df_global['Q21']=df_global['Q21'].replace(dict_format)

df_global['Q24'] = df_global['Q24'].replace(-1, np.nan)
df_global['Q24']=df_global['Q24'].replace(dict_format)



shift_matrix = pd.crosstab(df_global['Q24'], df_global['Q21'])
new_columns = pd.MultiIndex.from_product([['MOUB'], shift_matrix.columns])

# Assign the new MultiIndex to the DataFrame
shift_matrix.columns = new_columns


# shift_matrix = shift_matrix.rename_axis('Q24', axis='index')
# shift_matrix = shift_matrix.rename_axis('Q21', axis='columns')


inside_append_dataframe_with_blank_rows(output_path, shift_matrix)








# Count:Age wise brands 




Error_n = ['Table 3 : Count:Age wise brands ']

df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)




df_global=df.copy()
df_global.reset_index(drop=True, inplace=True)
df_global['Q6'] = pd.to_numeric(df_global['Q6'], errors='coerce')
def categorize_scores(score):
    
    if 21<=score <= 25:
        return '21-25 Yrs'
    elif 26 <= score <= 30:
        return '26-30 Yrs'
    elif 31<=score <= 35:
        return '31-35 Yrs'
    elif 36<=score<=40:
        return '36-40 Yrs'
    elif 41<=score<=45:
        return '41-45 Yrs'
    elif 46<=score<=50:
        return '46-50 Yrs'
    else:
        return 'Other'
    

df_global['Age'] = df_global['Q6'].apply(categorize_scores) 






df_dictformation = pd.read_excel(r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Data input\BHS 22-Oct-24_DataMap.xlsx','MOUB')



# Join two columns into a dictionary
# Assuming 'Column1' and 'Column2' are the names of the columns you want to join
dict_format = dict(zip(df_dictformation['Column'], df_dictformation['Brand']))




df_global['Q21'] = df_global['Q21'].replace(dict_format)

df_modified = df_global.loc[:,['Q21','Age']]

crosstab_result = pd.crosstab(df_modified['Age'], df_modified['Q21'])



# dynamic_headers = [f'brands{i}' for i in range(1, len(crosstab_result.columns) + 1)]

count1=pd.DataFrame(crosstab_result )
count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


inside_append_dataframe_with_blank_rows(output_path, count1)











# Count:SEC wise brands


Error_n = ['Table4 : Count:SEC wise brands']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()

input_string1=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48,49,50,51,52,53,54,55,56,57,58,59,60,61,62,63,64,65,66,67,68,69,70,71,72,73,74,75,76,77,78,79,80,81,82,83,84,85,86,87,88,89,90,91,92,93,94,95,96,97,98,99,100,101]
string="American Club ,Cavanders Gold Rich Taste ,Chancellor ,Charminar Filter ,Charms Special Blue ,Classic ,Editions Trio ,Flake Special Filter ,Focus Mint  ,Gold Flake Kings ,Gold Flake Premium ,Gold Flake Premium Neo Smart ,Gold Flake Indie Mint ,Gold Flake Special ,Gold Flake Special Mint ,Gold Flake Super Star ,India King ,Marlboro KSFT ,Marlboro Advance Compact ,Marlboro Pocket Filter ,Navy Cut Fruit ,Navy Cut Virginia Filter ,Red & White Select ,Red & White Select,Regent ,Regent Black ,Regent Cool ,Silk Cut Blue ,Silk Cut Filter ,Silk Cut Virginia ,Special Blues ,Special Red Longs ,Special Red Signature ,Stellar Cool Blast  ,T3 White  ,Total Refresh ,Total Royal Twist ,Total Spearmint ,Wave Cool Mint ,Wills Navy Cut Filter ,Wills Navy Cut,American Club Mint ,American Club Smash ,American Fruit ,Benson & Hedges  ,Berkeley ,Berkely ,Blue Charms ,Blue Charms ,Bristol ,Cavander Gold  ,Chancellor ,Charminar Plains ,Charms Regular Filter ,Classic Connect ,Classic Ice Burst ,Club One ,Duke ,Editions Active Mint ,Editions Ice Fruit ,Editions Spark ,Flake Excel ,Flake Liberty ,Flake Mint Switch ,Flake Nova ,Flake White ,Gold Flake Century ,Gold Flake Filter ,Gold Flake Filter ,Gold Flake Kings ,Gold Flake Kings Lights ,Gold Flake Kings SLK ,Golden Gold Flake ,Kingston ,Marlboro Clove ,Marlboro Fuse Beyond ,NATIONAL GOLD FLAKE ,Navy Cut Deluxe Filter ,Originals ,Panama ,Panama Filter ,Panama Filter ,Player's Fruit ,Player's Mint ,Royal ,Scissors Menthol  ,Classic ,Stellar Slims Define ,Stellar Slims Shift ,Total Active Mint ,Wave Fruity ,Wave Mint ,Will Flake Premium Filter ,Zaffran ,American Club Clove Magik ,Classic AlphaTec ,Classic Double Burst ,Classic Verve ,Stellar Define Pan ,Stellar Shift Duos ,Wills Insignia "
input_string2 = string.split(',')  
# Remove leading and trailing spaces from each item
input_string2 = [item.strip() for item in input_string2]
dictionary = dict(zip(input_string1, input_string2))
df_global['Q21'] = df_global['Q21'].replace(dictionary)

input_string1=[1,2,3,4]
input_string2=['SEC A','SEC B','SEC C','SEC D']
dictionary = dict(zip(input_string1, input_string2))
df_global['SEC_Q'] = df_global['SEC_Q'].replace(dictionary)


df_modified = df_global.loc[:,['SEC_Q','Q21']]
crosstab_result = pd.crosstab(df_modified['SEC_Q'], df_modified['Q21'])
count1=pd.DataFrame(crosstab_result )
count1['Total']=count1.sum(min_count=1,axis=1)


total=count1.sum()
# total.name='Total'
# count1=count1._append(total)
count1.loc['Total']=total


inside_append_dataframe_with_blank_rows(output_path, count1)














# Count:Price quoted for MOUB



Error_n = ['Table5 :Count:Price quoted for MOUB']
df_name= pd.DataFrame(Error_n) 
existing_df = pd.read_excel(output_path)
startrow = existing_df.shape[0] + 4
with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
    df_name.to_excel(writer, startrow=startrow, index=False, header=False)



df_global=df.copy()
df_global = df_global.dropna(subset=['Q28'])
df_global = df_global.dropna(subset=['Q29'])





# MOUB brand naming

df_dictformation = pd.read_excel(r'D:\Input\Cigarette\Brand Health Study(October 2024)\Input\Data input\BHS 22-Oct-24_DataMap.xlsx','MOUB')
dict_format = dict(zip(df_dictformation['Column'], df_dictformation['Brand']))
df_global['Q21'] = df_global['Q21'].replace(dict_format)


df_global1=df_global[df_global['Q28'] != -1]
df_global2=df_global[df_global['Q29'] != -1]

crosstab_result = pd.crosstab(df_global1['Q21'], df_global1['Q28'])
crosstab_result1 = pd.crosstab(df_global2['Q21'], df_global2['Q29'])


inside_append_dataframe_with_blank_rows(output_path, crosstab_result)

inside_append_dataframe_with_blank_rows(output_path, crosstab_result1)


















